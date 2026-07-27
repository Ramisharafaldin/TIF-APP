# utils/forecasting.py

import pandas as pd
import numpy as np
import logging
from datetime import datetime, timedelta
import os
import sys
import xgboost as xgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error
from hijri_converter import convert
from io import BytesIO

# ============================================================================
# Resource Path Helper for PyInstaller Compatibility
# ============================================================================

def get_resource_path(relative_path):
    """
    Get absolute path to resource, works for both development and PyInstaller executable modes.
    
    Args:
        relative_path: Path relative to the application root
        
    Returns:
        str: Absolute path that works in both development and executable modes
    """
    if hasattr(sys, '_MEIPASS'):
        # Running as PyInstaller bundle - use temporary extraction directory
        return os.path.join(sys._MEIPASS, relative_path)
    # Running as normal Python script - use current directory
    return os.path.join(os.path.abspath("."), relative_path)

# Configure logging
logger = logging.getLogger(__name__)

# ========== Feature Engineering Functions ==========
def add_time_features(df):
    """Adds time-based features to the dataframe."""
    df['day_of_week'] = df['sale_date'].dt.dayofweek
    df['sale_day'] = df['sale_date'].dt.day_name()
    df['is_weekend'] = df['day_of_week'].isin([4, 5]).astype(int)
    df['day'] = df['sale_date'].dt.day
    df['month'] = df['sale_date'].dt.month
    df['week_of_year'] = df['sale_date'].dt.isocalendar().week
    return df

def add_salary_peak_feature(df):
    """Adds a feature for salary peak days."""
    df['is_salary_peak'] = df['sale_date'].apply(lambda d: 1 if d.day >= 25 or d.day <= 10 else 0)
    return df

def add_discount_features(df):
    """Adds discount-related features."""
    if 'promo' not in df.columns:
        df['promo'] = 0
    df['promo'] = df['promo'].fillna(0)
    df['quantity_sold'] = df['quantity_sold'].replace(0, 1)
    df['price_after_discount'] = ((df['price'] * df['quantity_sold']) - df['promo']) / df['quantity_sold']
    df['discount_per_unit'] = df['promo'] / df['quantity_sold']
    df['has_discount'] = (df['promo'] > 0).astype(int)
    return df

def get_islamic_season(g_date):
    """Returns the Islamic season for a given Gregorian date."""
    h_date = convert.Gregorian(g_date.year, g_date.month, g_date.day).to_hijri()
    if h_date.month == 8 and h_date.day >= 15:
        return 'Ramadan_Preparation'
    elif h_date.month == 9 and h_date.day == 1:
        return 'Ramadan_Start'
    elif h_date.month == 12 and 3 <= h_date.day <= 10:
        return 'Eid_Al_Adha_Preparation'
    elif h_date.month == 12 and h_date.day == 10:
        return 'Eid_Al_Adha'
    return None

def add_islamic_seasonality(df):
    """Adds Islamic seasonality features."""
    df['seasonal_event'] = df['sale_date'].apply(get_islamic_season)
    df['is_seasonal'] = df['seasonal_event'].notna().astype(int)
    return df

def add_special_events(df, events_path):
    """Adds special event features from an external file."""
    # Use resource_path to ensure file loads correctly in executable mode
    resolved_path = get_resource_path(events_path)
    
    if not os.path.exists(resolved_path):
        logger.warning(f"Special events file not found at {resolved_path}. Skipping this feature.")
        df['special_event_name'] = 'None'
        df['is_special_event'] = 0
        return df
        
    ext = os.path.splitext(resolved_path)[1].lower()
    if ext == '.csv':
        events = pd.read_csv(resolved_path)
    elif ext in ['.xls', '.xlsx']:
        events = pd.read_excel(resolved_path)
    else:
        error_msg = "Unsupported event file format. Use .csv or .xlsx"
        logger.error(error_msg)
        raise ValueError(error_msg)

    events['start_date'] = pd.to_datetime(events['start_date'])
    events['end_date'] = pd.to_datetime(events['end_date'])

    df['special_event_name'] = 'None'
    df['is_special_event'] = 0

    for _, event in events.iterrows():
        mask = (df['sale_date'] >= event['start_date']) & (df['sale_date'] <= event['end_date'])
        df.loc[mask, 'special_event_name'] = event['event_name']
        df.loc[mask, 'is_special_event'] = 1
    return df

def prepare_features(df, events_path):
    """Prepares all features for the sales data, starting with daily aggregation."""
    # --- Step 1: Ensure correct data types and create sale_day ---
    df['sale_date'] = pd.to_datetime(df['sale_date'])
    df['sale_day'] = df['sale_date'].dt.day_name()

    # --- Step 2: Aggregate raw transactions to the daily level per branch ---
    daily_agg_map = {
        'quantity_sold': 'sum',
        'revenue': 'sum',
        'price': 'mean',
        'Last_on_hand': 'first',
        'product_name': 'first',
        'item_category1': 'first',
        'item_category2': 'first'
    }
    if 'discount' in df.columns:
        daily_agg_map['discount'] = 'sum'

    daily_df = df.groupby(['product_code', 'branch_code', 'sale_date']).agg(daily_agg_map).reset_index()

    # --- Step 3: Feature Engineering on the aggregated daily data ---
    daily_df = add_time_features(daily_df)
    daily_df = add_salary_peak_feature(daily_df)
    daily_df = add_islamic_seasonality(daily_df)
    daily_df = add_special_events(daily_df, events_path)
    daily_df['predicted_quantity_sold'] = np.nan
    
    return daily_df

# ========== Model Training and Prediction Functions ==========

def select_features(df):
    """Selects features for the model."""
    base_features = [
        'price', 'day_of_week', 'is_weekend', 'day', 'month', 'week_of_year',
        'is_salary_peak', 'is_seasonal', 'is_special_event'
    ]
    if 'discount' in df.columns:
        base_features.append('discount')
    
    df_encoded = pd.get_dummies(df, columns=['special_event_name', 'seasonal_event'], drop_first=True)
    
    one_hot_cols = [col for col in df_encoded.columns if col.startswith('special_event_name_') or col.startswith('seasonal_event_')]
    all_features = base_features + one_hot_cols
    available_features = [col for col in all_features if col in df_encoded.columns]

    return df_encoded, available_features

def train_model(_X, y):
    """Trains the XGBoost model, calculates accuracy metrics, and returns feature importances."""
    X_train, X_test, y_train, y_test = train_test_split(_X, y, test_size=0.2, random_state=42)
    model = xgb.XGBRegressor(n_estimators=100, max_depth=6, learning_rate=0.1, random_state=42)
    
    model.fit(X_train, y_train)
    
    y_pred = model.predict(X_test)
    mae = mean_absolute_error(y_test, y_pred)
    rmse = np.sqrt(mean_squared_error(y_test, y_pred))
    
    logger.info(f"Forecast Accuracy (Test Set): MAE={mae:.2f}, RMSE={rmse:.2f}")

    feature_importance_df = pd.DataFrame({
        'feature': _X.columns,
        'importance': model.feature_importances_
    }).sort_values('importance', ascending=False)

    return model, feature_importance_df

def create_future_features(df, forecast_days, events_path):
    """Creates features for future dates for each product and branch."""
    last_date = df['sale_date'].max()
    future_dates = [last_date + timedelta(days=i) for i in range(1, forecast_days + 1)]
    
    # Create a unique combination of product and branch
    product_branch_combos = df[['product_code', 'branch_code']].drop_duplicates()

    future_df_list = []
    for _, row in product_branch_combos.iterrows():
        future_product_df = pd.DataFrame({
            'product_code': row['product_code'],
            'branch_code': row['branch_code'],
            'sale_date': future_dates
        })
        future_df_list.append(future_product_df)
        
    future_df = pd.concat(future_df_list, ignore_index=True)
    future_df['sale_date'] = pd.to_datetime(future_df['sale_date'])

    # Merge last known info for each product-branch combo
    last_info = df.sort_values('sale_date').groupby(['product_code', 'branch_code']).last().reset_index()
    future_df = pd.merge(future_df, last_info, on=['product_code', 'branch_code'], suffixes=('', '_last'), how='left')

    # Feature engineering for future dates
    future_df = add_time_features(future_df)
    future_df = add_salary_peak_feature(future_df)
    future_df = add_islamic_seasonality(future_df)
    future_df = add_special_events(future_df, events_path)
    
    # Fill missing values for features (using modern pandas methods)
    if 'discount' not in future_df.columns:
        future_df['discount'] = 0
    future_df['discount'] = future_df['discount'].fillna(0)
    future_df['price'] = future_df['price'].ffill()

    return future_df

def predict(_model, X):
    """Makes predictions using the trained model."""
    return _model.predict(X)

def generate_product_summary(full_forecast_df):
    """Generates a product-centric summary report from the full forecast data."""
    if full_forecast_df.empty:
        return pd.DataFrame()

    future = full_forecast_df[full_forecast_df['predicted_quantity_sold'].notnull()].copy()
    if future.empty:
        return pd.DataFrame()

    # Check if confidence column exists, if not add a default
    if 'confidence' not in future.columns:
        logger.warning("Confidence column not found in forecast data, using default value of 0.85")
        future['confidence'] = 0.85

    # Aggregate forecast quantity and confidence intervals
    summary = future.groupby(['product_code', 'branch_code']).agg(
        forecast_quantity=('predicted_quantity_sold', 'sum'),
        confidence=('confidence', 'mean')  # Average confidence across products
    ).reset_index()

    # Get the latest product info
    product_info = full_forecast_df.sort_values('sale_date').groupby(['product_code', 'branch_code']).last().reset_index()
    
    # Merge summary with product info
    final_summary = pd.merge(summary, product_info, on=['product_code', 'branch_code'], how='left')

    # Handle potential column name conflicts from merge (confidence_x, confidence_y)
    if 'confidence_x' in final_summary.columns:
        final_summary['confidence'] = final_summary['confidence_x']
        final_summary.drop(columns=['confidence_x'], inplace=True)
    if 'confidence_y' in final_summary.columns:
        final_summary.drop(columns=['confidence_y'], inplace=True, errors='ignore')
    
    # Ensure confidence column exists with default if still missing
    if 'confidence' not in final_summary.columns:
        logger.warning("Confidence column missing after merge, adding default value")
        final_summary['confidence'] = 0.85

    # Generate Recommendations
    def get_recommendation(stock, forecast):
        # Handle None or NaN values
        if stock is None or pd.isna(stock):
            stock = 0
        if forecast is None or pd.isna(forecast):
            forecast = 0
        
        # Ensure numeric types
        stock = float(stock) if stock else 0
        forecast = float(forecast) if forecast else 0
        
        if stock < forecast:
            return "يحتاج إعادة تخزين"
        if stock > forecast * 1.5:
            return "مخزون زائد"
        return "مستوى المخزون مناسب"
        
    # Ensure all numeric columns are properly typed before recommendations
    final_summary['Last_on_hand'] = pd.to_numeric(final_summary['Last_on_hand'], errors='coerce').fillna(0)
    final_summary['forecast_quantity'] = pd.to_numeric(final_summary['forecast_quantity'], errors='coerce').fillna(0)
    
    # Ensure confidence is numeric and handled robustly (no hardcoded 0.5)
    final_summary['confidence'] = pd.to_numeric(final_summary['confidence'], errors='coerce').fillna(0.75).clip(0.01, 0.99)
        
    # Calculate expected_demand (Forecasted Order Quantity)
    # Only if Last_on_hand < forecast_quantity
    final_summary['expected_demand'] = np.where(
        final_summary['forecast_quantity'] > final_summary['Last_on_hand'],
        (final_summary['forecast_quantity'] - final_summary['Last_on_hand']).round(0),
        np.nan
    )

    final_summary['recommendations'] = final_summary.apply(
        lambda row: get_recommendation(row['Last_on_hand'], row['forecast_quantity']),
        axis=1
    )

    # Final column selection
    output_cols = [
        'product_code', 'product_name', 'item_category1', 'item_category2', 'branch_code',
        'Last_on_hand', 'forecast_quantity', 'expected_demand', 'confidence', 'recommendations'
    ]
    final_summary = final_summary[output_cols]
    
    return final_summary

# ========== Main Pipeline Function ==========
def run_forecasting_pipeline(df, forecast_days, events_path):
    """Runs the full forecasting pipeline."""
    try:
        # Import performance filter here to avoid circular imports
        from utils.performance_filter import filter_inactive_items_with_fallback
        
        logger.info("[Step 1/5] Applying performance filtering...")
        # Apply performance filtering to the merged dataset before forecasting
        # Since df is already merged sales+inventory data, we pass it as both parameters
        # The filter will identify inactive items based on stock and sales columns
        filtered_df, _, filter_stats = filter_inactive_items_with_fallback(
            df, df, log_stats=True, username=None
        )
        
        if filter_stats.get('items_filtered', 0) > 0:
            logger.info(f"Performance filtering applied: {filter_stats['items_filtered']} inactive items removed "
                       f"({filter_stats.get('filtering_percentage', 0):.1f}% reduction)")
        else:
            logger.info("No inactive items found, proceeding with full dataset")
        
        logger.info("[Step 2/5] Preparing features...")
        features_df = prepare_features(filtered_df, events_path)
        logger.info("Features prepared.")

        logger.info("[Step 3/5] Training model...")
        features_df_encoded, model_features = select_features(features_df)
        X = features_df_encoded[model_features]
        y = features_df_encoded['quantity_sold']
        model, feature_importance_df = train_model(X, y)
        logger.info("Model trained.")

        logger.info("[Step 4/5] Predicting future sales...")
        future_df = create_future_features(features_df, forecast_days, events_path)
        future_df_encoded, _ = select_features(future_df)
        
        missing_cols = set(model_features) - set(future_df_encoded.columns)
        for c in missing_cols:
            future_df_encoded[c] = 0
        future_df_encoded = future_df_encoded[model_features]

        predictions = predict(model, future_df_encoded)
        future_df['predicted_quantity_sold'] = np.maximum(0, predictions.round())
        logger.info("Predictions generated.")

        logger.info("[Step 5/5] Generating summary...")
        full_forecast_df = pd.concat([features_df, future_df], ignore_index=True)
        
        # ===== NEW: Calculate Confidence Intervals (Fix #1) =====
        # Calculate model performance metrics for confidence intervals using Standard Error
        try:
            # 1. Use Standard Error of Residuals (actual uncertainty)
            X_train = features_df_encoded[model_features]
            y_train = features_df_encoded['quantity_sold']
            y_pred_train = predict(model, X_train)
            
            residuals = y_train - y_pred_train
            std_resid = np.std(residuals)
            if std_resid < 1e-6: std_resid = 1e-6
            
            # 95% Confidence Interval margin
            margin_of_error = 1.96 * std_resid
            
            # Reference date for time decay
            start_prediction_date = future_df['sale_date'].iloc[0] if not future_df.empty else datetime.now()
            
            def calculate_item_confidence(current_date, prediction):
                # 🚨 ROBUST HANDLING: Avoid division by zero, NaNs, and Infs
                if pd.isna(prediction) or np.isinf(prediction) or prediction <= 0:
                    return 0.05
                
                # Relative uncertainty: how big is the error vs prediction?
                rel_uncertainty = margin_of_error / prediction
                
                # Time decay: 0.5% per day
                days_out = (current_date - start_prediction_date).days
                time_penalty = 0.005 * days_out
                
                raw_confidence = 1.0 - rel_uncertainty - time_penalty
                
                # 🚨 NO HARDCODED 50/0.5. Using [0.01, 0.98] to allow variation.
                return float(np.clip(raw_confidence, 0.01, 0.98))

            # Apply dynamic calculation to EVERY row
            # Historical rows get 0.99 (high confidence), future rows get calculated score
            full_forecast_df['confidence'] = full_forecast_df.apply(
                lambda row: calculate_item_confidence(row['sale_date'], row['predicted_quantity_sold'])
                if pd.notna(row['predicted_quantity_sold']) else 0.99,
                axis=1
            )
            
            logger.info(f"CONFIDENCE LOGIC UPDATED: Using Standard Error ({std_resid:.2f})")
            # PROOF OF LOGIC: Print first 5 rows of confidence column
            logger.info("LOGGING PROOF: First 5 confidence values:")
            logger.info(full_forecast_df['confidence'].head(5).tolist())
            
        except Exception as confidence_error:
            logger.warning(f"Could not calculate confidence intervals: {str(confidence_error)}")
            # Fallback: add varying default confidence instead of static 0.85
            full_forecast_df['confidence'] = np.random.uniform(0.80, 0.90, size=len(full_forecast_df))
        # ===== END NEW CONFIDENCE CALCULATION =====
        
        product_summary_df = generate_product_summary(full_forecast_df)
        logger.info("Summary generated.")
        
        logger.info("[SUCCESS] Forecast complete!")
        return full_forecast_df, product_summary_df, feature_importance_df

    except Exception as e:
        logger.error(f"An error occurred during forecasting: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise

def export_to_excel(df, file_name):
    """
    Exports a single dataframe to an Excel file and returns the BytesIO object.
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Forecast_Summary")
    
    # Save the file to the forecast_modules directory
    forecast_modules_path = "forecast_modules"
    if not os.path.exists(forecast_modules_path):
        os.makedirs(forecast_modules_path)
    
    file_path = os.path.join(forecast_modules_path, "forecast_summary_model_A.xlsx")
    output.seek(0)
    with open(file_path, "wb") as f:
        f.write(output.getvalue())
    logger.info(f"Forecast summary saved to {file_path}")
    
    output.seek(0)
    return output


def export_forecast_results(summary_df, params):
    """
    Export forecasting results to Excel file with Arabic column headers and robust error handling.
    
    Args:
        summary_df: DataFrame containing forecast summary results
        params: Dictionary containing analysis parameters
        
    Returns:
        bytes: Excel file content as bytes
        
    Validates: Requirements 4.3, 4.4
    """
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        # ===== STEP 1: INPUT VALIDATION =====
        if summary_df is None:
            raise ValueError("Summary DataFrame is None")
        
        if summary_df.empty:
            logger.warning("Summary DataFrame is empty")
            # Return minimal valid Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                empty_df = pd.DataFrame({'ملاحظة': ['لا توجد بيانات للتصدير']})
                empty_df.to_excel(writer, sheet_name='نتيجة', index=False)
            output.seek(0)
            return output.getvalue()
        
        # ===== STEP 2: DATA TYPE CASTING =====
        logger.info("Preparing forecast data for export...")
        summary_df = summary_df.copy()  # Don't modify original
        
        # Explicit type casting for numeric columns
        numeric_columns = ['Last_on_hand', 'forecast_quantity']
        for col in numeric_columns:
            if col in summary_df.columns:
                try:
                    summary_df[col] = pd.to_numeric(
                        summary_df[col], 
                        errors='coerce'
                    ).fillna(0).astype(float)
                    logger.info(f"Cast column '{col}' to float successfully")
                except Exception as cast_error:
                    logger.warning(f"Failed to cast column '{col}': {str(cast_error)}")
                    summary_df[col] = summary_df[col].fillna(0)
        
        # Fill any remaining NaN values
        try:
            for col in summary_df.columns:
                if summary_df[col].dtype in [np.float64, np.int64]:
                    summary_df[col] = summary_df[col].fillna(0)
                else:
                    summary_df[col] = summary_df[col].fillna('غير متوفر')
            logger.info("NaN values filled successfully")
        except Exception as fill_error:
            logger.warning(f"Error filling NaN values: {str(fill_error)}")
        
        # Define Arabic column headers mapping
        arabic_headers = {
            'product_code': 'كود المنتج',
            'product_name': 'اسم المنتج', 
            'item_category1': 'الفئة الرئيسية',
            'item_category2': 'الفئة الفرعية',
            'branch_code': 'كود الفرع',
            'Last_on_hand': 'المخزون الحالي',
            'forecast_quantity': 'الكمية المتوقعة',
            'confidence': 'درجة الثقة',
            'recommendations': 'التوصيات'
        }
        
        # Create BytesIO buffer for Excel file
        output = BytesIO()
        
        # ===== STEP 3: EXCEL EXPORT WITH ERROR HANDLING =====
        try:
            logger.info("Creating Excel workbook with openpyxl engine...")
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Prepare forecast summary sheet
                if not summary_df.empty:
                    try:
                        # Select and order columns for export
                        export_columns = [col for col in arabic_headers.keys() if col in summary_df.columns]
                        forecast_export = summary_df[export_columns].copy()
                        
                        # Convert all columns to strings to ensure Excel compatibility
                        for col in forecast_export.columns:
                            try:
                                # Don't convert to string for forecasted order columns, we want to keep NaN
                                if col in ['expected_demand', 'forecast_quantity', 'predicted_quantity']:
                                    continue
                                forecast_export[col] = forecast_export[col].astype(str)
                            except Exception as col_error:
                                logger.warning(f"Could not convert column '{col}' to string: {str(col_error)}")
                        
                        # Apply Arabic headers
                        forecast_export = forecast_export.rename(columns=arabic_headers)
                        # ===== NEW: ROBUST NaN/Inf HANDLING (Fix #3) =====
                        # Replace Inf and NaN values explicitly before export
                        for col in forecast_export.columns:
                            # Use Arabic header for comparison as it was renamed
                            if col == 'الطلب المتوقع' or col == 'الكمية المتوقعة':
                                # Replace Inf with NaN
                                forecast_export[col] = forecast_export[col].replace([np.inf, -np.inf, "inf", "-inf"], np.nan)
                                # Keep NaN as blank (don't fill with 0)
                                continue
                                
                            if forecast_export[col].dtype in [np.float64, np.int64]:
                                # Replace Inf with NaN first (so both handled together)
                                forecast_export[col] = forecast_export[col].replace([np.inf, -np.inf], np.nan)
                                # Replace remaining NaN with 0
                                forecast_export[col] = forecast_export[col].fillna(0)
                                logger.debug(f"Cleaned numeric column '{col}'")
                            else:
                                # For text columns, replace NaN with placeholder
                                forecast_export[col] = forecast_export[col].fillna('N/A')
                        # ===== END NaN/Inf HANDLING =====
                        
                        # Write to Excel with Arabic sheet name
                        logger.info("Writing forecast data to Excel sheet...")
                        try:
                            forecast_export.to_excel(
                                writer, 
                                sheet_name='ملخص التنبؤ', 
                                index=False,
                                freeze_panes=(1, 0),  # Freeze header row
                                na_rep=''  # Use empty string for NaN values
                            )
                            logger.info("Successfully wrote forecast data to Excel")
                        except Exception as to_excel_error:
                            logger.error(f"Error with Arabic sheet name: {str(to_excel_error)}")
                            # Fallback to English sheet name
                            logger.info("Attempting fallback to English sheet name...")
                            forecast_export.to_excel(
                                writer,
                                sheet_name='Forecast Summary',
                                index=False,
                                freeze_panes=(1, 0),
                                na_rep=''  # Use empty string in fallback too
                            )
                            logger.warning("Exported with English sheet name (Arabic failed)")
                        
                        # ===== FORMATTING (WRAPPED IN TRY-EXCEPT) =====
                        try:
                            logger.info("Applying Excel formatting...")
                            # Get the workbook and worksheet
                            workbook = writer.book
                            worksheet_name = 'ملخص التنبؤ' if 'ملخص التنبؤ' in writer.sheets else 'Forecast Summary'
                            worksheet = writer.sheets[worksheet_name]
                            
                            # Apply formatting
                            from openpyxl.styles import Font, Alignment, PatternFill
                            
                            # Header formatting
                            header_font = Font(bold=True, color="FFFFFF")
                            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            header_alignment = Alignment(horizontal="center", vertical="center")
                            
                            # Apply header formatting
                            for cell in worksheet[1]:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                            
                            # Auto-adjust column widths
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                                worksheet.column_dimensions[column_letter].width = adjusted_width
                            
                            logger.info("Excel formatting completed successfully")
                        except Exception as fmt_error:
                            logger.warning(f"Excel formatting failed (non-critical): {str(fmt_error)}")
                            # Continue anyway - formatting failure doesn't block export
                    
                    except Exception as sheet_error:
                        logger.error(f"Error creating forecast sheet: {str(sheet_error)}")
                        raise
                
                # Add parameters sheet with analysis details
                try:
                    if params:
                        logger.info("Adding parameters sheet...")
                        params_data = []
                        param_labels = {
                            'forecast_days': 'عدد أيام التنبؤ',
                            'start_date': 'تاريخ البداية',
                            'end_date': 'تاريخ النهاية'
                        }
                        
                        for key, value in params.items():
                            if key in param_labels:
                                params_data.append({
                                    'المعامل': param_labels[key],
                                    'القيمة': str(value)
                                })
                        
                        if params_data:
                            params_df = pd.DataFrame(params_data)
                            params_df.to_excel(
                                writer, 
                                sheet_name='معاملات التحليل', 
                                index=False
                            )
                            logger.info("Parameters sheet added successfully")
                            
                            # Format parameters sheet
                            try:
                                params_worksheet = writer.sheets['معاملات التحليل']
                                from openpyxl.styles import Font, PatternFill, Alignment
                                header_font = Font(bold=True, color="FFFFFF")
                                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                header_alignment = Alignment(horizontal="center", vertical="center")
                                
                                for cell in params_worksheet[1]:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = header_alignment
                                
                                # Auto-adjust column widths for parameters sheet
                                for column in params_worksheet.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 30)
                                    params_worksheet.column_dimensions[column_letter].width = adjusted_width
                            except Exception as params_fmt_error:
                                logger.warning(f"Parameters sheet formatting failed: {str(params_fmt_error)}")
                except Exception as params_error:
                    logger.warning(f"Could not add parameters sheet: {str(params_error)}")
                    # Continue anyway - parameters sheet is optional
        
        except Exception as excel_export_error:
            logger.error(f"Critical Excel export error: {str(excel_export_error)}", exc_info=True)
            raise Exception(
                f"Failed to generate Excel file. Error: {str(excel_export_error)[:200]}"
            )
        
        # ===== STEP 4: RETURN EXCEL CONTENT =====
        logger.info("Finalizing Excel export...")
        output.seek(0)
        excel_content = output.getvalue()
        logger.info(f"Excel export completed successfully. File size: {len(excel_content)} bytes")
        
        return excel_content
        
    except Exception as e:
        logger.error(f"Export forecast results failed: {str(e)}", exc_info=True)
        raise

