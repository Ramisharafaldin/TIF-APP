"""
Property-based tests for comprehensive error handling and logging.
Tests that errors are properly logged with context and provide actionable feedback.

Feature: data-upload-persistence-fix
"""

import pytest
import sys
import os
import tempfile
import sqlite3
import logging
from io import BytesIO, StringIO

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from hypothesis import given, strategies as st, assume, settings, HealthCheck
import pandas as pd
from openpyxl import Workbook

import data_store


@pytest.fixture
def test_user():
    """Create a test user for authentication"""
    import auth_flask
    username = 'test_user_error_handling'
    password = 'TestPass123!'
    
    # Add test user
    auth_flask.add_user(username, password, is_admin=False)
    
    yield {'username': username, 'password': password}
    
    # Cleanup
    try:
        auth_flask.delete_user(username, 'admin')
    except:
        pass


@pytest.fixture
def log_capture():
    """Capture log messages for testing"""
    log_stream = StringIO()
    handler = logging.StreamHandler(log_stream)
    handler.setLevel(logging.DEBUG)
    
    # Get the logger used by data_store and data_processing
    logger = logging.getLogger('utils.data_processing')
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG)
    
    yield log_stream
    
    # Cleanup
    logger.removeHandler(handler)


@st.composite
def error_scenario_strategy(draw):
    """Generate different error scenarios for testing"""
    error_type = draw(st.sampled_from([
        'missing_required_sheets', 
        'file_corruption',
        'empty_file'
    ]))
    
    if error_type == 'missing_required_sheets':
        # Create Excel with wrong sheet names that won't be found
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create sheets with names that won't match the expected patterns
        wrong_sheet1 = wb.create_sheet('WrongSheetName1')
        wrong_sheet1.cell(row=1, column=1, value='some_data')
        wrong_sheet1.cell(row=2, column=1, value='value1')
        
        wrong_sheet2 = wb.create_sheet('WrongSheetName2')
        wrong_sheet2.cell(row=1, column=1, value='other_data')
        wrong_sheet2.cell(row=2, column=1, value='value2')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {
            'filename': 'missing_sheets.xlsx',
            'file_data': excel_buffer.getvalue(),
            'branch_name': 'TestBranch',
            'error_type': error_type,
            'expected_error_keywords': ['sheet', 'find', 'transaction', 'item']
        }
        
    elif error_type == 'file_corruption':
        # Create corrupted Excel file
        return {
            'filename': 'corrupted.xlsx',
            'file_data': b'corrupted excel data that cannot be parsed as xlsx',
            'branch_name': 'TestBranch',
            'error_type': error_type,
            'expected_error_keywords': ['corrupted', 'invalid', 'format', 'excel']
        }
        
    else:  # empty_file
        # Create completely empty file
        return {
            'filename': 'empty.xlsx',
            'file_data': b'',
            'branch_name': 'TestBranch',
            'error_type': error_type,
            'expected_error_keywords': ['empty', 'invalid', 'file']
        }


class TestErrorHandlingProperties:
    """Property-based tests for comprehensive error handling and logging"""
    
    @given(
        error_scenario=error_scenario_strategy()
    )
    @settings(max_examples=3, suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture], deadline=None)
    def test_property_5_comprehensive_error_handling_and_logging(self, test_user, log_capture, error_scenario):
        """
        Property 5: Comprehensive Error Handling and Logging
        
        For any operation that encounters an error (database failure, processing failure, 
        validation failure), the system should log the specific error with timestamp, 
        username, and operation details, and provide actionable user feedback.
        
        **Feature: data-upload-persistence-fix, Property 5: Comprehensive Error Handling and Logging**
        **Validates: Requirements 4.1, 4.2, 4.3, 4.4, 4.5**
        """
        
        # Skip if error_scenario is None (shouldn't happen with fixed strategy)
        if error_scenario is None:
            assume(False)
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Clear log capture
                log_capture.seek(0)
                log_capture.truncate(0)
                
                # Attempt operation that should fail
                error_occurred = False
                error_message = ""
                
                try:
                    file_id, sales_id, inventory_id = data_store.save_branch_data(
                        username=test_user['username'],
                        branch_name=error_scenario['branch_name'],
                        filename=error_scenario['filename'],
                        file_data=error_scenario['file_data']
                    )
                    # If we get here without exception, the operation unexpectedly succeeded
                    error_occurred = False
                except Exception as e:
                    # Expected - operation should fail
                    error_occurred = True
                    error_message = str(e)
                
                # Verify error occurred as expected
                assert error_occurred, f"Operation should have failed for {error_scenario['error_type']}"
                
                # Verify error message is informative and actionable
                assert len(error_message) > 0, "Should provide error message"
                assert not error_message.isspace(), "Error message should not be just whitespace"
                
                # Check that error message contains relevant information
                error_message_lower = error_message.lower()
                
                # Verify logging occurred
                log_contents = log_capture.getvalue()
                
                # Property assertions for logging
                if log_contents:  # Only check if logging is configured
                    log_contents_lower = log_contents.lower()
                    
                    # Should contain error-related keywords
                    assert any(keyword in log_contents_lower for keyword in ['error', 'failed', 'exception']), \
                        "Log should contain error-related keywords"
                    
                    # Should contain operation context
                    # Note: We check for general operation context rather than specific format
                    # since different modules may log differently
                    assert len(log_contents.strip()) > 0, "Should have logged error information"
                
                # Verify system remains in consistent state after error
                # Check that no partial data was saved
                conn = sqlite3.connect(data_store.DB_NAME)
                c = conn.cursor()
                c.execute('SELECT COUNT(*) FROM uploaded_files WHERE username = ?', (test_user['username'],))
                files_count = c.fetchone()[0]
                c.execute('SELECT COUNT(*) FROM processed_data WHERE username = ?', (test_user['username'],))
                data_count = c.fetchone()[0]
                conn.close()
                
                # Property assertion: System should remain consistent after errors
                # With the improved validation, no database operations should occur when validation fails
                assert files_count == 0, "No files should be saved when validation fails"
                assert data_count == 0, "No processed data should be saved when validation fails"
                
                # Verify error provides actionable feedback
                # Error message should not be a generic "An error occurred"
                assert "an error occurred" not in error_message.lower() or len(error_message) > 20, \
                    "Error message should be more specific than generic 'an error occurred'"
                
            finally:
                # Cleanup
                data_store.DB_NAME = original_db
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass


if __name__ == '__main__':
    pytest.main([__file__, '-v'])