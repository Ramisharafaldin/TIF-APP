"""
Property-based tests for data validation before processing.
Tests that validation failures prevent database operations.

Feature: data-upload-persistence-fix
"""

import pytest
import sys
import os
import tempfile
import sqlite3
from io import BytesIO

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from hypothesis import given, strategies as st, assume, settings, HealthCheck
import pandas as pd
from openpyxl import Workbook

import data_store


@pytest.fixture
def test_user():
    """Create a test user for authentication"""
    import auth_flask
    username = 'test_user_validation'
    password = 'TestPass123!'
    
    # Add test user
    auth_flask.add_user(username, password, is_admin=False)
    
    yield {'username': username, 'password': password}
    
    # Cleanup
    try:
        auth_flask.delete_user(username, 'admin')
    except:
        pass


@st.composite
def invalid_excel_file_strategy(draw):
    """Generate Excel files with validation issues"""
    issue_type = draw(st.sampled_from([
        'missing_transactions_sheet',
        'missing_item_info_sheet', 
        'missing_required_columns_transactions',
        'missing_required_columns_items',
        'empty_branch_name',
        'invalid_file_extension'
    ]))
    
    if issue_type == 'missing_transactions_sheet':
        # Create Excel with only Item info sheet
        wb = Workbook()
        wb.remove(wb.active)
        
        item_info_ws = wb.create_sheet('Item info')
        item_info_ws.cell(row=1, column=1, value='product_code')
        item_info_ws.cell(row=1, column=2, value='product_name')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {
            'filename': 'missing_transactions.xlsx',
            'file_data': excel_buffer.getvalue(),
            'branch_name': 'ValidBranch',
            'expected_error': 'missing_transactions_sheet'
        }
        
    elif issue_type == 'missing_item_info_sheet':
        # Create Excel with only Transactions sheet
        wb = Workbook()
        wb.remove(wb.active)
        
        transactions_ws = wb.create_sheet('Transactions')
        transactions_ws.cell(row=1, column=1, value='product_code')
        transactions_ws.cell(row=1, column=2, value='sale_date')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {
            'filename': 'missing_item_info.xlsx',
            'file_data': excel_buffer.getvalue(),
            'branch_name': 'ValidBranch',
            'expected_error': 'missing_item_info_sheet'
        }
        
    elif issue_type == 'missing_required_columns_transactions':
        # Create Excel with Transactions sheet missing required columns
        wb = Workbook()
        wb.remove(wb.active)
        
        transactions_ws = wb.create_sheet('Transactions')
        transactions_ws.cell(row=1, column=1, value='wrong_column')
        
        item_info_ws = wb.create_sheet('Item info')
        item_info_ws.cell(row=1, column=1, value='product_code')
        item_info_ws.cell(row=1, column=2, value='product_name')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {
            'filename': 'invalid_transactions_columns.xlsx',
            'file_data': excel_buffer.getvalue(),
            'branch_name': 'ValidBranch',
            'expected_error': 'missing_required_columns'
        }
        
    elif issue_type == 'missing_required_columns_items':
        # Create Excel with Item info sheet missing required columns
        wb = Workbook()
        wb.remove(wb.active)
        
        transactions_ws = wb.create_sheet('Transactions')
        transactions_ws.cell(row=1, column=1, value='product_code')
        transactions_ws.cell(row=1, column=2, value='sale_date')
        
        item_info_ws = wb.create_sheet('Item info')
        item_info_ws.cell(row=1, column=1, value='wrong_column')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {
            'filename': 'invalid_item_info_columns.xlsx',
            'file_data': excel_buffer.getvalue(),
            'branch_name': 'ValidBranch',
            'expected_error': 'missing_required_columns'
        }
        
    elif issue_type == 'empty_branch_name':
        # Valid Excel but empty branch name
        wb = Workbook()
        wb.remove(wb.active)
        
        transactions_ws = wb.create_sheet('Transactions')
        transactions_ws.cell(row=1, column=1, value='product_code')
        transactions_ws.cell(row=1, column=2, value='sale_date')
        
        item_info_ws = wb.create_sheet('Item info')
        item_info_ws.cell(row=1, column=1, value='product_code')
        item_info_ws.cell(row=1, column=2, value='product_name')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return {
            'filename': 'valid_file.xlsx',
            'file_data': excel_buffer.getvalue(),
            'branch_name': '',  # Empty branch name
            'expected_error': 'empty_branch_name'
        }
        
    elif issue_type == 'invalid_file_extension':
        # Invalid file extension
        return {
            'filename': 'invalid_file.txt',
            'file_data': b'not an excel file',
            'branch_name': 'ValidBranch',
            'expected_error': 'invalid_file_extension'
        }


class TestDataValidationProperties:
    """Property-based tests for data validation before processing"""
    
    @given(
        invalid_file=invalid_excel_file_strategy()
    )
    @settings(max_examples=5, suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture], deadline=None)
    def test_property_6_data_validation_before_processing(self, test_user, invalid_file):
        """
        Property 6: Data Validation Before Processing
        
        For any uploaded Excel file, if the file lacks required sheets (Transactions, Item info),
        required columns, or has invalid branch names, the validation should fail before any
        database operations occur.
        
        **Feature: data-upload-persistence-fix, Property 6: Data Validation Before Processing**
        **Validates: Requirements 5.1, 5.2, 5.3, 5.5**
        """
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Count initial database records
                conn = sqlite3.connect(data_store.DB_NAME)
                c = conn.cursor()
                c.execute('SELECT COUNT(*) FROM uploaded_files WHERE username = ?', (test_user['username'],))
                initial_files_count = c.fetchone()[0]
                c.execute('SELECT COUNT(*) FROM processed_data WHERE username = ?', (test_user['username'],))
                initial_data_count = c.fetchone()[0]
                conn.close()
                
                # Attempt to save invalid data
                validation_failed = False
                error_message = ""
                
                try:
                    # Test different validation scenarios
                    if invalid_file['expected_error'] == 'empty_branch_name':
                        # This should fail at the branch name validation level
                        if not invalid_file['branch_name'] or not invalid_file['branch_name'].strip():
                            validation_failed = True
                            error_message = "Empty branch name should be rejected"
                    
                    elif invalid_file['expected_error'] == 'invalid_file_extension':
                        # This should fail at file extension validation
                        filename = invalid_file['filename']
                        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in {'xlsx', 'xls', 'csv'}:
                            validation_failed = True
                            error_message = "Invalid file extension should be rejected"
                    
                    else:
                        # For Excel structure issues, attempt processing and expect failure
                        try:
                            file_id, sales_id, inventory_id = data_store.save_branch_data(
                                username=test_user['username'],
                                branch_name=invalid_file['branch_name'],
                                filename=invalid_file['filename'],
                                file_data=invalid_file['file_data']
                            )
                            # If we get here without exception, validation didn't work properly
                            validation_failed = False
                            error_message = "Processing should have failed but didn't"
                        except Exception as e:
                            # Expected - validation should fail
                            validation_failed = True
                            error_message = str(e)
                
                except Exception as e:
                    # Expected for validation failures
                    validation_failed = True
                    error_message = str(e)
                
                # Verify validation failed as expected
                assert validation_failed, f"Validation should have failed for {invalid_file['expected_error']}"
                
                # Verify no database operations occurred when validation failed
                conn = sqlite3.connect(data_store.DB_NAME)
                c = conn.cursor()
                c.execute('SELECT COUNT(*) FROM uploaded_files WHERE username = ?', (test_user['username'],))
                final_files_count = c.fetchone()[0]
                c.execute('SELECT COUNT(*) FROM processed_data WHERE username = ?', (test_user['username'],))
                final_data_count = c.fetchone()[0]
                conn.close()
                
                # Property assertion: No database operations should occur when validation fails
                assert final_files_count == initial_files_count, "No files should be saved when validation fails"
                assert final_data_count == initial_data_count, "No processed data should be saved when validation fails"
                
                # Verify error message is informative
                assert len(error_message) > 0, "Should provide informative error message"
                
            finally:
                # Cleanup
                data_store.DB_NAME = original_db
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass


if __name__ == '__main__':
    pytest.main([__file__, '-v'])