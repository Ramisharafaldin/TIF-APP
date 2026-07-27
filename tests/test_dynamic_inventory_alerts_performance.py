"""
Performance tests for Dynamic Inventory Alerts feature.
Tests alert generation with large datasets, concurrent access scenarios, and query performance.

Feature: dynamic-inventory-alerts
Task: 7.1 Write performance tests
Requirements: 6.1, 6.4
"""

import pytest
import sys
import os
import tempfile
import sqlite3
import time
import threading
import concurrent.futures
from io import BytesIO
from datetime import datetime, timedelta

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from openpyxl import Workbook

import data_store
from utils import alert_service


@pytest.fixture
def test_user():
    """Create a test user for authentication"""
    import auth_flask
    username = 'test_user_perf'
    password = 'TestPass123!'
    
    # Add test user
    auth_flask.add_user(username, password, is_admin=False)
    
    yield {'username': username, 'password': password}
    
    # Cleanup
    try:
        auth_flask.delete_user(username, 'admin')
    except:
        pass


def create_large_inventory_dataset(num_products=1000, branch_name="Branch1"):
    """
    Create a large inventory dataset for performance testing.
    
    Args:
        num_products: Number of products to generate
        branch_name: Branch name to use for all products
        
    Returns:
        DataFrame with inventory data
    """
    import random
    
    data = []
    
    for i in range(num_products):
        # Create stock levels with realistic distribution
        # 10% out of stock, 20% very low, 30% low, 25% reorder, 15% normal
        rand = random.random()
        if rand < 0.10:
            stock_level = 0  # Out of stock
        elif rand < 0.30:
            stock_level = random.randint(1, 5)  # Very low
        elif rand < 0.60:
            stock_level = random.randint(6, 15)  # Low
        elif rand < 0.85:
            stock_level = random.randint(16, 25)  # Reorder
        else:
            stock_level = random.randint(26, 100)  # Normal stock
        
        data.append({
            'product_code': f"P{i+1:06d}",
            'product_name': f"Product {i+1} - Test Item for Performance Testing",
            'branch_code': branch_name,
            'Last_on_hand': stock_level,
            'supplier_name': f"Supplier {random.randint(1, 20)}",
            'item_category1': f"Category {random.randint(1, 10)}",
            'inventory_value': round(random.uniform(10.0, 1000.0), 2)
        })
    
    return pd.DataFrame(data)


def create_excel_with_large_dataset(inventory_df):
    """
    Create Excel file with large inventory dataset.
    
    Args:
        inventory_df: DataFrame with inventory data
        
    Returns:
        BytesIO buffer with Excel file data
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create minimal sales sheet (required for data processing)
    sales_ws = wb.create_sheet('Transactions')
    sales_headers = ['product_code', 'sale_date', 'quantity', 'unit_price', 'revenue']
    for col, header in enumerate(sales_headers, 1):
        sales_ws.cell(row=1, column=col, value=header)
    
    # Add minimal sales data for first 100 products (to keep file size manageable)
    for row, (_, item) in enumerate(inventory_df.head(100).iterrows(), 2):
        sales_ws.cell(row=row, column=1, value=item['product_code'])
        sales_ws.cell(row=row, column=2, value='2024-01-01')
        sales_ws.cell(row=row, column=3, value=1)
        sales_ws.cell(row=row, column=4, value=10.0)
        sales_ws.cell(row=row, column=5, value=10.0)
    
    # Create inventory sheet
    inventory_ws = wb.create_sheet('Item info')
    headers = list(inventory_df.columns)
    for col, header in enumerate(headers, 1):
        inventory_ws.cell(row=1, column=col, value=header)
    
    for row, (_, item) in enumerate(inventory_df.iterrows(), 2):
        for col, header in enumerate(headers, 1):
            inventory_ws.cell(row=row, column=col, value=item[header])
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


class TestDynamicInventoryAlertsPerformance:
    """Performance tests for dynamic inventory alerts"""
    
    def test_large_dataset_performance(self, test_user):
        """
        Test alert generation with large datasets.
        
        Requirements: 6.1 - System should complete processing within 2 seconds for datasets up to 10,000 products
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Test with different dataset sizes
                dataset_sizes = [100, 500, 1000, 2000, 5000]
                
                for size in dataset_sizes:
                    print(f"\nTesting with {size} products...")
                    
                    # Create large inventory dataset
                    inventory_df = create_large_inventory_dataset(num_products=size, branch_name="Branch1")
                    
                    # Create Excel file
                    excel_buffer = create_excel_with_large_dataset(inventory_df)
                    
                    # Upload data and measure time
                    upload_start = time.time()
                    file_id, sales_id, inventory_id = data_store.save_branch_data(
                        username=test_user['username'],
                        branch_name="Branch1",
                        filename=f'large_dataset_{size}.xlsx',
                        file_data=excel_buffer.getvalue()
                    )
                    upload_time = time.time() - upload_start
                    
                    print(f"Upload time for {size} products: {upload_time:.2f}s")
                    
                    # Generate alerts and measure time
                    alert_start = time.time()
                    alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter="Branch1",
                        limit=1000  # Get many alerts for testing
                    )
                    alert_time = time.time() - alert_start
                    
                    print(f"Alert generation time for {size} products: {alert_time:.2f}s")
                    print(f"Generated {len(alerts)} alerts")
                    
                    # Performance assertions
                    if size <= 1000:
                        # For smaller datasets, should be very fast
                        assert alert_time < 1.0, \
                            f"Alert generation for {size} products should complete within 1 second, took {alert_time:.2f}s"
                    elif size <= 5000:
                        # For medium datasets, should meet the 2-second requirement
                        assert alert_time < 2.0, \
                            f"Alert generation for {size} products should complete within 2 seconds, took {alert_time:.2f}s"
                    
                    # Verify alerts are properly generated
                    assert isinstance(alerts, list), "Should return list of alerts"
                    assert len(alerts) > 0, f"Should generate some alerts for {size} products"
                    
                    # Verify alert quality is maintained despite large dataset
                    for alert in alerts[:10]:  # Check first 10 alerts
                        assert alert.product_code is not None, "Alert should have valid product_code"
                        assert alert.current_stock >= 0, "Alert should have valid stock level"
                        assert alert.branch_code == "Branch1", "Alert should have correct branch"
                        assert 1 <= alert.priority <= 4, "Alert should have valid priority"
                    
                    # Clear data for next iteration
                    data_store.clear_user_data(test_user['username'])
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    def test_concurrent_access_scenarios(self, test_user):
        """
        Test concurrent access scenarios.
        
        Requirements: 6.4 - System should handle concurrent dashboard requests without performance degradation
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Create test dataset
                inventory_df = create_large_inventory_dataset(num_products=1000, branch_name="Branch1")
                excel_buffer = create_excel_with_large_dataset(inventory_df)
                
                # Upload test data
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username=test_user['username'],
                    branch_name="Branch1",
                    filename='concurrent_test.xlsx',
                    file_data=excel_buffer.getvalue()
                )
                
                # Test concurrent alert generation
                def generate_alerts_worker(worker_id):
                    """Worker function for concurrent alert generation"""
                    start_time = time.time()
                    try:
                        alerts = alert_service.generate_inventory_alerts(
                            username=test_user['username'],
                            branch_filter="Branch1",
                            limit=10
                        )
                        end_time = time.time()
                        
                        return {
                            'worker_id': worker_id,
                            'success': True,
                            'alerts_count': len(alerts),
                            'duration': end_time - start_time,
                            'error': None
                        }
                    except Exception as e:
                        end_time = time.time()
                        return {
                            'worker_id': worker_id,
                            'success': False,
                            'alerts_count': 0,
                            'duration': end_time - start_time,
                            'error': str(e)
                        }
                
                # Test with different numbers of concurrent workers
                concurrent_levels = [2, 5, 10]
                
                for num_workers in concurrent_levels:
                    print(f"\nTesting with {num_workers} concurrent workers...")
                    
                    # Run concurrent alert generation
                    start_time = time.time()
                    
                    with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
                        # Submit all tasks
                        futures = [executor.submit(generate_alerts_worker, i) for i in range(num_workers)]
                        
                        # Wait for all tasks to complete
                        results = [future.result() for future in concurrent.futures.as_completed(futures)]
                    
                    total_time = time.time() - start_time
                    
                    print(f"Total time for {num_workers} concurrent requests: {total_time:.2f}s")
                    
                    # Analyze results
                    successful_requests = [r for r in results if r['success']]
                    failed_requests = [r for r in results if not r['success']]
                    
                    print(f"Successful requests: {len(successful_requests)}/{num_workers}")
                    print(f"Failed requests: {len(failed_requests)}")
                    
                    if successful_requests:
                        avg_duration = sum(r['duration'] for r in successful_requests) / len(successful_requests)
                        max_duration = max(r['duration'] for r in successful_requests)
                        min_duration = min(r['duration'] for r in successful_requests)
                        
                        print(f"Average request duration: {avg_duration:.2f}s")
                        print(f"Max request duration: {max_duration:.2f}s")
                        print(f"Min request duration: {min_duration:.2f}s")
                    
                    # Performance assertions
                    assert len(successful_requests) >= num_workers * 0.8, \
                        f"At least 80% of concurrent requests should succeed, got {len(successful_requests)}/{num_workers}"
                    
                    if successful_requests:
                        # No request should take more than 5 seconds
                        assert max_duration < 5.0, \
                            f"No concurrent request should take more than 5 seconds, max was {max_duration:.2f}s"
                        
                        # Average duration should be reasonable
                        assert avg_duration < 3.0, \
                            f"Average concurrent request duration should be under 3 seconds, was {avg_duration:.2f}s"
                        
                        # All successful requests should return consistent results
                        alert_counts = [r['alerts_count'] for r in successful_requests]
                        assert len(set(alert_counts)) <= 2, \
                            f"Concurrent requests should return consistent alert counts, got {set(alert_counts)}"
                    
                    # Print any errors for debugging
                    for result in failed_requests:
                        print(f"Worker {result['worker_id']} failed: {result['error']}")
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    def test_query_performance_measurement(self, test_user):
        """
        Measure query performance for alert generation.
        
        Requirements: 6.1, 6.4 - Measure and validate query performance
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Create test dataset with multiple branches
                inventory_df = create_large_inventory_dataset(num_products=2000, branch_name="Branch1")
                excel_buffer = create_excel_with_large_dataset(inventory_df)
                
                # Upload test data
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username=test_user['username'],
                    branch_name="Branch1",
                    filename='query_perf_test.xlsx',
                    file_data=excel_buffer.getvalue()
                )
                
                # Test different query scenarios
                test_scenarios = [
                    {'name': 'All branches, limit 10', 'branch_filter': None, 'limit': 10},
                    {'name': 'All branches, limit 50', 'branch_filter': None, 'limit': 50},
                    {'name': 'All branches, limit 100', 'branch_filter': None, 'limit': 100},
                    {'name': 'Single branch, limit 10', 'branch_filter': 'Branch1', 'limit': 10},
                    {'name': 'Single branch, limit 50', 'branch_filter': 'Branch1', 'limit': 50},
                    {'name': 'Single branch, no limit', 'branch_filter': 'Branch1', 'limit': 1000},
                ]
                
                performance_results = []
                
                for scenario in test_scenarios:
                    print(f"\nTesting scenario: {scenario['name']}")
                    
                    # Run multiple iterations to get average performance
                    durations = []
                    alert_counts = []
                    
                    for iteration in range(5):  # 5 iterations per scenario
                        start_time = time.time()
                        
                        alerts = alert_service.generate_inventory_alerts(
                            username=test_user['username'],
                            branch_filter=scenario['branch_filter'],
                            limit=scenario['limit']
                        )
                        
                        duration = time.time() - start_time
                        durations.append(duration)
                        alert_counts.append(len(alerts))
                    
                    # Calculate statistics
                    avg_duration = sum(durations) / len(durations)
                    max_duration = max(durations)
                    min_duration = min(durations)
                    avg_alerts = sum(alert_counts) / len(alert_counts)
                    
                    result = {
                        'scenario': scenario['name'],
                        'avg_duration': avg_duration,
                        'max_duration': max_duration,
                        'min_duration': min_duration,
                        'avg_alerts': avg_alerts,
                        'branch_filter': scenario['branch_filter'],
                        'limit': scenario['limit']
                    }
                    
                    performance_results.append(result)
                    
                    print(f"Average duration: {avg_duration:.3f}s")
                    print(f"Max duration: {max_duration:.3f}s")
                    print(f"Min duration: {min_duration:.3f}s")
                    print(f"Average alerts: {avg_alerts:.1f}")
                    
                    # Performance assertions
                    assert avg_duration < 2.0, \
                        f"Average query duration should be under 2 seconds for '{scenario['name']}', was {avg_duration:.3f}s"
                    
                    assert max_duration < 3.0, \
                        f"Max query duration should be under 3 seconds for '{scenario['name']}', was {max_duration:.3f}s"
                    
                    # Verify consistent results across iterations
                    unique_alert_counts = set(alert_counts)
                    assert len(unique_alert_counts) <= 2, \
                        f"Alert counts should be consistent across iterations for '{scenario['name']}', got {unique_alert_counts}"
                
                # Test cache performance
                print("\nTesting cache performance...")
                
                # First call (no cache)
                start_time = time.time()
                alerts_first = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter='Branch1',
                    limit=50
                )
                first_call_duration = time.time() - start_time
                
                # Second call (should use cache)
                start_time = time.time()
                alerts_second = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter='Branch1',
                    limit=50
                )
                second_call_duration = time.time() - start_time
                
                print(f"First call (no cache): {first_call_duration:.3f}s")
                print(f"Second call (cached): {second_call_duration:.3f}s")
                
                # Cache should make second call faster
                assert second_call_duration <= first_call_duration, \
                    f"Cached call should be faster or equal: first={first_call_duration:.3f}s, second={second_call_duration:.3f}s"
                
                # Results should be identical
                assert len(alerts_first) == len(alerts_second), \
                    "Cached results should have same number of alerts"
                
                for i, (alert1, alert2) in enumerate(zip(alerts_first, alerts_second)):
                    assert alert1.product_code == alert2.product_code, \
                        f"Cached alert {i} should have same product_code"
                    assert alert1.current_stock == alert2.current_stock, \
                        f"Cached alert {i} should have same stock level"
                
                # Print performance summary
                print("\n=== Performance Summary ===")
                for result in performance_results:
                    print(f"{result['scenario']}: {result['avg_duration']:.3f}s avg, {result['avg_alerts']:.0f} alerts")
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    def test_memory_usage_with_large_datasets(self, test_user):
        """
        Test memory usage with large datasets to ensure no memory leaks.
        
        Requirements: 6.1, 6.4 - Ensure system handles large datasets without memory issues
        """
        import psutil
        import gc
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Get initial memory usage
                process = psutil.Process()
                initial_memory = process.memory_info().rss / 1024 / 1024  # MB
                
                print(f"Initial memory usage: {initial_memory:.1f} MB")
                
                # Test with progressively larger datasets
                dataset_sizes = [500, 1000, 1500, 2000]
                memory_measurements = []
                
                for size in dataset_sizes:
                    print(f"\nTesting memory usage with {size} products...")
                    
                    # Create large dataset
                    inventory_df = create_large_inventory_dataset(num_products=size, branch_name="Branch1")
                    excel_buffer = create_excel_with_large_dataset(inventory_df)
                    
                    # Upload data
                    file_id, sales_id, inventory_id = data_store.save_branch_data(
                        username=test_user['username'],
                        branch_name="Branch1",
                        filename=f'memory_test_{size}.xlsx',
                        file_data=excel_buffer.getvalue()
                    )
                    
                    # Generate alerts multiple times to test for memory leaks
                    for iteration in range(10):
                        alerts = alert_service.generate_inventory_alerts(
                            username=test_user['username'],
                            branch_filter="Branch1",
                            limit=100
                        )
                        
                        # Force garbage collection
                        gc.collect()
                    
                    # Measure memory after processing
                    current_memory = process.memory_info().rss / 1024 / 1024  # MB
                    memory_increase = current_memory - initial_memory
                    
                    memory_measurements.append({
                        'dataset_size': size,
                        'memory_mb': current_memory,
                        'memory_increase_mb': memory_increase
                    })
                    
                    print(f"Memory usage after {size} products: {current_memory:.1f} MB (+{memory_increase:.1f} MB)")
                    
                    # Memory usage should not grow excessively
                    assert memory_increase < 200, \
                        f"Memory increase should be under 200 MB for {size} products, was {memory_increase:.1f} MB"
                    
                    # Clear data for next iteration
                    data_store.clear_user_data(test_user['username'])
                    
                    # Force garbage collection
                    gc.collect()
                
                # Check for memory leaks - memory should not grow linearly with dataset size
                if len(memory_measurements) >= 2:
                    # Calculate memory growth rate
                    first_measurement = memory_measurements[0]
                    last_measurement = memory_measurements[-1]
                    
                    size_ratio = last_measurement['dataset_size'] / first_measurement['dataset_size']
                    memory_ratio = last_measurement['memory_increase_mb'] / max(first_measurement['memory_increase_mb'], 1)
                    
                    print(f"\nMemory growth analysis:")
                    print(f"Dataset size ratio: {size_ratio:.1f}x")
                    print(f"Memory increase ratio: {memory_ratio:.1f}x")
                    
                    # Memory growth should be sub-linear (not proportional to dataset size)
                    assert memory_ratio < size_ratio * 0.8, \
                        f"Memory growth should be sub-linear: size ratio {size_ratio:.1f}x, memory ratio {memory_ratio:.1f}x"
                
                print("\n=== Memory Usage Summary ===")
                for measurement in memory_measurements:
                    print(f"{measurement['dataset_size']} products: {measurement['memory_mb']:.1f} MB "
                          f"(+{measurement['memory_increase_mb']:.1f} MB)")
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass


if __name__ == "__main__":
    # Run performance tests directly
    pytest.main([__file__, "-v", "-s"])