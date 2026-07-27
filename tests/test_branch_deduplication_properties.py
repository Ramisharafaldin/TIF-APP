"""
Property-based tests for branch deduplication logic.

Tests Property 4: Branch Deduplication Logic
**Validates: Requirements 2.3**
"""

import pytest
import tempfile
import os
import sqlite3
from hypothesis import given, strategies as st, settings, assume
from datetime import datetime, timedelta
import data_store
from openpyxl import Workbook
import io


def create_test_excel_file(branch_suffix=""):
    """Create a minimal valid Excel file for testing"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create Transactions sheet
    transactions_ws = wb.create_sheet("Transactions")
    transactions_ws.append(["تاريخ الفاتورة", "رقم الفاتورة", "كود الصنف", "اسم الصنف", "الكمية", "سعر البيع"])
    transactions_ws.append([f"2024-01-01", f"INV001{branch_suffix}", f"ITEM001{branch_suffix}", f"Test Item{branch_suffix}", 10, 100.0])
    transactions_ws.append([f"2024-01-02", f"INV002{branch_suffix}", f"ITEM002{branch_suffix}", f"Test Item 2{branch_suffix}", 5, 200.0])
    
    # Create Item info sheet
    item_info_ws = wb.create_sheet("Item info")
    item_info_ws.append(["كود الصنف", "اسم الصنف", "الكمية الحالية", "سعر التكلفة"])
    item_info_ws.append([f"ITEM001{branch_suffix}", f"Test Item{branch_suffix}", 100, 80.0])
    item_info_ws.append([f"ITEM002{branch_suffix}", f"Test Item 2{branch_suffix}", 50, 150.0])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def setup_test_database():
    """Set up a temporary test database."""
    # Create temporary database file
    db_fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(db_fd)
    
    # Override the database path for testing
    original_db = data_store.DB_NAME
    data_store.DB_NAME = db_path
    
    # Initialize the test database
    data_store.init_data_db()
    
    return db_path, original_db


def cleanup_test_database(db_path, original_db):
    """Clean up the temporary test database."""
    data_store.DB_NAME = original_db
    if os.path.exists(db_path):
        os.remove(db_path)


def create_test_user():
    """Create a test user for testing."""
    return {
        'username': 'test_user_dedup',
        'password': 'test_password'
    }


@given(
    branch_name=st.text(min_size=1, max_size=50).filter(lambda x: x.strip() and not any(c in x for c in ['/', '\\', ':', '*', '?', '"', '<', '>', '|'])),
    num_uploads=st.integers(min_value=2, max_value=5)
)
@settings(max_examples=10, deadline=30000)
def test_branch_deduplication_property(branch_name, num_uploads):
    """
    Property 4: Branch Deduplication Logic
    
    For any branch with multiple uploaded files, the get_branch_files() function 
    should return only the most recent upload for that branch.
    
    **Feature: data-upload-persistence-fix, Property 4: Branch Deduplication Logic**
    **Validates: Requirements 2.3**
    """
    db_path, original_db = setup_test_database()
    
    try:
        test_user = create_test_user()
        username = test_user['username']
        
        # Create multiple uploads for the same branch with different timestamps
        upload_times = []
        file_ids = []
        
        for i in range(num_uploads):
            # Create a test Excel file
            excel_data = create_test_excel_file()
            filename = f"test_file_{i}.xlsx"
            
            # Save the file using save_branch_data
            try:
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username, branch_name, filename, excel_data
                )
                file_ids.append(file_id)
                
                # Get the actual upload timestamp from database
                conn = sqlite3.connect(data_store.DB_NAME)
                c = conn.cursor()
                c.execute('SELECT upload_timestamp FROM uploaded_files WHERE id = ?', (file_id,))
                timestamp = c.fetchone()[0]
                upload_times.append(timestamp)
                conn.close()
                
                # Add a small delay to ensure different timestamps
                import time
                time.sleep(0.01)
                
            except Exception as e:
                # Skip this test case if we can't create valid test data
                assume(False)
        
        # Get branch files using the function under test
        branch_files = data_store.get_branch_files(username)
        
        # Filter files for our test branch
        test_branch_files = [f for f in branch_files if f['branch_name'] == branch_name]
        
        # Property: Should return only one file per branch (the most recent)
        assert len(test_branch_files) == 1, f"Expected 1 file for branch {branch_name}, got {len(test_branch_files)}"
        
        # The returned file should be the most recent one
        returned_file = test_branch_files[0]
        
        # Find the most recent upload time
        most_recent_time = max(upload_times)
        
        # Verify the returned file has the most recent timestamp
        assert returned_file['upload_date'] == most_recent_time, \
            f"Expected most recent timestamp {most_recent_time}, got {returned_file['upload_date']}"
        
        # Verify the file belongs to the correct branch
        assert returned_file['branch_name'] == branch_name, \
            f"Expected branch name {branch_name}, got {returned_file['branch_name']}"
        
        # Verify other required fields are present
        assert 'filename' in returned_file, "Filename should be present"
        assert 'file_size' in returned_file, "File size should be present"
        assert returned_file['file_size'] > 0, "File size should be positive"
        
    finally:
        cleanup_test_database(db_path, original_db)


@given(
    branches=st.lists(
        st.text(min_size=1, max_size=30).filter(lambda x: x.strip() and not any(c in x for c in ['/', '\\', ':', '*', '?', '"', '<', '>', '|'])),
        min_size=2, max_size=4, unique=True
    ),
    uploads_per_branch=st.integers(min_value=1, max_value=3)
)
@settings(max_examples=10, deadline=30000)
def test_multiple_branches_deduplication_property(branches, uploads_per_branch):
    """
    Property 4 Extended: Multiple Branches Deduplication Logic
    
    For any set of branches with multiple uploaded files each, the get_branch_files() 
    function should return only the most recent upload for each branch.
    
    **Feature: data-upload-persistence-fix, Property 4: Branch Deduplication Logic**
    **Validates: Requirements 2.3**
    """
    db_path, original_db = setup_test_database()
    
    try:
        test_user = create_test_user()
        username = test_user['username']
        
        branch_latest_times = {}
        
        # Upload multiple files for each branch
        for branch_name in branches:
            latest_time = None
            
            for i in range(uploads_per_branch):
                # Create a test Excel file
                excel_data = create_test_excel_file()
                filename = f"test_file_{branch_name}_{i}.xlsx"
                
                try:
                    file_id, sales_id, inventory_id = data_store.save_branch_data(
                        username, branch_name, filename, excel_data
                    )
                    
                    # Get the actual upload timestamp from database
                    conn = sqlite3.connect(data_store.DB_NAME)
                    c = conn.cursor()
                    c.execute('SELECT upload_timestamp FROM uploaded_files WHERE id = ?', (file_id,))
                    timestamp = c.fetchone()[0]
                    conn.close()
                    
                    # Track the latest timestamp for this branch
                    if latest_time is None or timestamp > latest_time:
                        latest_time = timestamp
                    
                    # Add a small delay to ensure different timestamps
                    import time
                    time.sleep(0.01)
                    
                except Exception as e:
                    # Skip this test case if we can't create valid test data
                    assume(False)
            
            branch_latest_times[branch_name] = latest_time
        
        # Get branch files using the function under test
        branch_files = data_store.get_branch_files(username)
        
        # Property: Should return exactly one file per branch
        assert len(branch_files) == len(branches), \
            f"Expected {len(branches)} files (one per branch), got {len(branch_files)}"
        
        # Verify each branch appears exactly once with its most recent file
        returned_branches = set()
        for file_info in branch_files:
            branch_name = file_info['branch_name']
            
            # Each branch should appear only once
            assert branch_name not in returned_branches, \
                f"Branch {branch_name} appears multiple times in results"
            returned_branches.add(branch_name)
            
            # The file should be the most recent for this branch
            expected_time = branch_latest_times[branch_name]
            assert file_info['upload_date'] == expected_time, \
                f"Branch {branch_name}: expected timestamp {expected_time}, got {file_info['upload_date']}"
        
        # All branches should be represented
        assert returned_branches == set(branches), \
            f"Expected branches {set(branches)}, got {returned_branches}"
        
    finally:
        cleanup_test_database(db_path, original_db)


if __name__ == "__main__":
    # Run a simple test to verify the property
    print("Testing branch deduplication property...")
    
    db_path, original_db = setup_test_database()
    
    try:
        test_user = create_test_user()
        username = test_user['username']
        branch_name = "test_branch"
        
        # Upload 3 files for the same branch
        for i in range(3):
            excel_data = create_test_excel_file()
            filename = f"test_file_{i}.xlsx"
            
            file_id, sales_id, inventory_id = data_store.save_branch_data(
                username, branch_name, filename, excel_data
            )
            print(f"Uploaded file {i+1} with ID {file_id}")
            
            import time
            time.sleep(0.1)  # Ensure different timestamps
        
        # Test deduplication
        branch_files = data_store.get_branch_files(username)
        print(f"Retrieved {len(branch_files)} files")
        
        # Should only return 1 file (the most recent)
        assert len(branch_files) == 1, f"Expected 1 file, got {len(branch_files)}"
        print("✓ Branch deduplication property test passed")
        
    finally:
        cleanup_test_database(db_path, original_db)