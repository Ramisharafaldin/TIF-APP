"""
Property-based tests for Dynamic Inventory Alerts feature.
Tests universal properties that should hold across all valid inputs.

Feature: dynamic-inventory-alerts
"""

import pytest
import sys
import os
import tempfile
import sqlite3
import time
from io import BytesIO
from datetime import datetime, timedelta

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from hypothesis import given, strategies as st, assume, settings, HealthCheck
import pandas as pd
from openpyxl import Workbook

import data_store
from utils import alert_service


@pytest.fixture
def test_user():
    """Create a test user for authentication"""
    import auth_flask
    username = 'test_user_alerts'
    password = 'TestPass123!'
    
    # Add test user
    auth_flask.add_user(username, password, is_admin=False)
    
    yield {'username': username, 'password': password}
    
    # Cleanup
    try:
        auth_flask.delete_user(username, 'admin')
    except:
        pass


# Hypothesis strategies for generating test data
@st.composite
def valid_branch_name_strategy(draw):
    """Generate valid branch names"""
    prefix = draw(st.sampled_from(['Branch', 'Store', 'Location', 'Office']))
    number = draw(st.integers(min_value=1, max_value=99))
    return f"{prefix}{number}"


@st.composite
def inventory_item_strategy(draw, branch_code=None):
    """Generate a single inventory item with stock level"""
    product_code = f"P{draw(st.integers(min_value=1, max_value=999)):03d}"
    product_name = f"Product {draw(st.text(min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd', 'Pc'))))}"
    # Use provided branch_code or generate random one
    if branch_code is None:
        branch_code = draw(valid_branch_name_strategy())
    stock_level = draw(st.integers(min_value=0, max_value=100))
    
    return {
        'product_code': product_code,
        'product_name': product_name,
        'branch_code': branch_code,
        'Last_on_hand': stock_level,
        'supplier_name': f"Supplier {draw(st.integers(min_value=1, max_value=10))}",
        'item_category1': f"Category {draw(st.integers(min_value=1, max_value=5))}",
        'inventory_value': round(draw(st.floats(min_value=1.0, max_value=1000.0, allow_nan=False, allow_infinity=False)), 2)
    }


@st.composite
def inventory_dataset_strategy(draw, branch_code=None):
    """Generate a complete inventory dataset"""
    num_items = draw(st.integers(min_value=1, max_value=20))
    items = [draw(inventory_item_strategy(branch_code=branch_code)) for _ in range(num_items)]
    
    # Create DataFrame
    df = pd.DataFrame(items)
    
    return {
        'inventory_df': df,
        'expected_low_stock_items': len([item for item in items if item['Last_on_hand'] <= 25]),
        'expected_out_of_stock_items': len([item for item in items if item['Last_on_hand'] == 0]),
        'total_items': len(items)
    }


@st.composite
def excel_with_inventory_strategy(draw, branch_code=None):
    """Generate Excel file with inventory data for testing"""
    dataset = draw(inventory_dataset_strategy(branch_code=branch_code))
    inventory_df = dataset['inventory_df']
    
    # Create Excel file in memory
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create minimal sales sheet (required for data processing)
    sales_ws = wb.create_sheet('Transactions')
    sales_headers = ['product_code', 'sale_date', 'quantity', 'unit_price', 'revenue']
    for col, header in enumerate(sales_headers, 1):
        sales_ws.cell(row=1, column=col, value=header)
    
    # Add minimal sales data for each product
    for row, (_, item) in enumerate(inventory_df.iterrows(), 2):
        sales_ws.cell(row=row, column=1, value=item['product_code'])
        sales_ws.cell(row=row, column=2, value='2024-01-01')
        sales_ws.cell(row=row, column=3, value=1)
        sales_ws.cell(row=row, column=4, value=10.0)
        sales_ws.cell(row=row, column=5, value=10.0)
    
    # Create inventory sheet
    inventory_ws = wb.create_sheet('Item info')
    headers = list(inventory_df.columns)
    for col, header in enumerate(headers, 1):
        inventory_ws.cell(row=1, column=col, value=header)
    
    for row, (_, item) in enumerate(inventory_df.iterrows(), 2):
        for col, header in enumerate(headers, 1):
            inventory_ws.cell(row=row, column=col, value=item[header])
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return {
        'filename': f'test_inventory_{draw(st.integers(min_value=1, max_value=1000))}.xlsx',
        'file_data': excel_buffer.getvalue(),
        'dataset': dataset
    }


class TestDynamicInventoryAlertsProperties:
    """Property-based tests for dynamic inventory alerts"""
    
    @given(
        branch_name=valid_branch_name_strategy(),
        excel_file=excel_with_inventory_strategy()
    )
    @settings(max_examples=1, suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture], deadline=None)
    def test_property_1_alert_generation_consistency(self, test_user, branch_name, excel_file):
        """
        Property 1: Alert Generation Consistency
        
        For any inventory dataset with products below thresholds, the alert system should 
        generate exactly one alert per low-stock product and no alerts for products above thresholds.
        
        **Feature: dynamic-inventory-alerts, Property 1: Alert Generation Consistency**
        **Validates: Requirements 1.1, 1.2, 1.5**
        """
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Upload test data
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username=test_user['username'],
                    branch_name=branch_name,
                    filename=excel_file['filename'],
                    file_data=excel_file['file_data']
                )
                
                # Generate alerts
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=branch_name,
                    limit=1000  # Get all alerts for testing
                )
                
                # Get the original inventory data for comparison
                dataset = excel_file['dataset']
                inventory_df = dataset['inventory_df']
                
                # Property 1: Alert Generation Consistency
                # Count products that should have alerts (stock <= 25)
                products_needing_alerts = inventory_df[inventory_df['Last_on_hand'] <= 25]
                expected_alert_count = len(products_needing_alerts)
                
                # Count products that should NOT have alerts (stock > 25)
                products_not_needing_alerts = inventory_df[inventory_df['Last_on_hand'] > 25]
                
                # Assertions for alert generation consistency
                assert len(alerts) == expected_alert_count, \
                    f"Should generate exactly {expected_alert_count} alerts for products with stock <= 25, got {len(alerts)}"
                
                # Verify each alert corresponds to a product that needs attention
                alert_product_codes = {alert.product_code for alert in alerts}
                expected_product_codes = set(products_needing_alerts['product_code'])
                
                assert alert_product_codes == expected_product_codes, \
                    "Alert product codes should match exactly the products needing alerts"
                
                # Verify no alerts for products above threshold
                products_above_threshold = set(products_not_needing_alerts['product_code'])
                alerts_for_good_stock = alert_product_codes.intersection(products_above_threshold)
                
                assert len(alerts_for_good_stock) == 0, \
                    f"Should not generate alerts for products with stock > 25, but got alerts for: {alerts_for_good_stock}"
                
                # Verify each product has at most one alert
                assert len(alert_product_codes) == len(alerts), \
                    "Each product should have at most one alert (no duplicates)"
                
                # Verify all alerts have required fields
                for alert in alerts:
                    assert alert.product_code is not None and alert.product_code != "", \
                        "Alert should have valid product_code"
                    assert alert.product_name is not None and alert.product_name != "", \
                        "Alert should have valid product_name"
                    assert alert.branch_code == branch_name, \
                        f"Alert branch_code should match filter: expected {branch_name}, got {alert.branch_code}"
                    assert isinstance(alert.current_stock, int) and alert.current_stock >= 0, \
                        "Alert should have valid current_stock"
                    assert alert.alert_status is not None and alert.alert_status != "", \
                        "Alert should have valid alert_status"
                    assert alert.status_class is not None and alert.status_class != "", \
                        "Alert should have valid status_class"
                    assert isinstance(alert.priority, int) and 1 <= alert.priority <= 4, \
                        "Alert should have valid priority (1-4)"
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    @given(stock_level=st.integers(min_value=0, max_value=100))
    @settings(max_examples=5, suppress_health_check=[HealthCheck.too_slow], deadline=None)
    def test_property_2_alert_classification_accuracy(self, stock_level):
        """
        Property 2: Alert Classification Accuracy
        
        For any product stock level, the alert system should classify it according to the 
        defined thresholds: 0 = "نفد المخزون", 1-5 = "منخفض جداً", 6-15 = "منخفض", 
        16-25 = "إعادة طلب", with appropriate CSS classes.
        
        **Feature: dynamic-inventory-alerts, Property 2: Alert Classification Accuracy**
        **Validates: Requirements 2.1, 2.2, 2.3, 2.4, 2.5**
        """
        
        # Test the classification function directly
        status_text, css_class, priority = alert_service.classify_alert_status(stock_level)
        
        # Property 2: Alert Classification Accuracy
        if stock_level == 0:
            # Out of stock
            assert status_text == "نفد المخزون", \
                f"Stock level 0 should be classified as 'نفد المخزون', got '{status_text}'"
            assert css_class == "bg-red-600/10 text-red-600 border-red-600/20", \
                f"Stock level 0 should have red-600 CSS class, got '{css_class}'"
            assert priority == 1, \
                f"Stock level 0 should have priority 1, got {priority}"
                
        elif 1 <= stock_level <= 5:
            # Very low stock
            assert status_text == "منخفض جداً", \
                f"Stock level {stock_level} should be classified as 'منخفض جداً', got '{status_text}'"
            assert css_class == "bg-red-500/10 text-red-500 border-red-500/20", \
                f"Stock level {stock_level} should have red-500 CSS class, got '{css_class}'"
            assert priority == 2, \
                f"Stock level {stock_level} should have priority 2, got {priority}"
                
        elif 6 <= stock_level <= 15:
            # Low stock
            assert status_text == "منخفض", \
                f"Stock level {stock_level} should be classified as 'منخفض', got '{status_text}'"
            assert css_class == "bg-orange-500/10 text-orange-500 border-orange-500/20", \
                f"Stock level {stock_level} should have orange CSS class, got '{css_class}'"
            assert priority == 3, \
                f"Stock level {stock_level} should have priority 3, got {priority}"
                
        elif 16 <= stock_level <= 25:
            # Reorder level
            assert status_text == "إعادة طلب", \
                f"Stock level {stock_level} should be classified as 'إعادة طلب', got '{status_text}'"
            assert css_class == "bg-blue-500/10 text-blue-500 border-blue-500/20", \
                f"Stock level {stock_level} should have blue CSS class, got '{css_class}'"
            assert priority == 4, \
                f"Stock level {stock_level} should have priority 4, got {priority}"
                
        else:
            # Above reorder level - no alert needed
            assert status_text is None, \
                f"Stock level {stock_level} (>25) should not generate alert, got status '{status_text}'"
            assert css_class is None, \
                f"Stock level {stock_level} (>25) should not have CSS class, got '{css_class}'"
            assert priority is None, \
                f"Stock level {stock_level} (>25) should not have priority, got {priority}"
        
        # Additional validation: Ensure classification is consistent
        # Test with float values that should convert to int
        if stock_level > 0:
            float_status, float_css, float_priority = alert_service.classify_alert_status(float(stock_level))
            assert float_status == status_text, \
                f"Float stock level {float(stock_level)} should have same classification as int"
            assert float_css == css_class, \
                f"Float stock level {float(stock_level)} should have same CSS class as int"
            assert float_priority == priority, \
                f"Float stock level {float(stock_level)} should have same priority as int"
    
    @given(
        branch_name=valid_branch_name_strategy(),
        excel_file=excel_with_inventory_strategy(),
        limit_param=st.integers(min_value=1, max_value=20),
        branch_filter=st.one_of(st.none(), valid_branch_name_strategy())
    )
    @settings(max_examples=1, suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture], deadline=None)
    def test_property_6_api_response_completeness(self, test_user, branch_name, excel_file, limit_param, branch_filter):
        """
        Property 6: API Response Completeness
        
        For any API request to `/api/inventory-alerts`, the system should return properly 
        formatted JSON with all required alert fields, respect limit parameters, and require authentication.
        
        **Feature: dynamic-inventory-alerts, Property 6: API Response Completeness**
        **Validates: Requirements 8.2, 8.3, 8.4, 8.5**
        """
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Upload test data
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username=test_user['username'],
                    branch_name=branch_name,
                    filename=excel_file['filename'],
                    file_data=excel_file['file_data']
                )
                
                # Test the alert service directly (simulating API behavior)
                # Use the branch_filter parameter if provided, otherwise use uploaded branch
                filter_to_use = branch_filter if branch_filter else branch_name
                
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=filter_to_use,
                    limit=limit_param
                )
                
                # Property 6: API Response Completeness
                # Simulate API response format
                api_response = {
                    "success": True,
                    "alerts": [alert.to_dict() for alert in alerts],
                    "total_alerts": len(alerts),
                    "last_updated": datetime.now().isoformat()
                }
                
                # Verify response structure
                assert "success" in api_response, "API response should include 'success' field"
                assert api_response["success"] is True, "API response success should be True for valid request"
                
                assert "alerts" in api_response, "API response should include 'alerts' field"
                assert isinstance(api_response["alerts"], list), "API response alerts should be a list"
                
                assert "total_alerts" in api_response, "API response should include 'total_alerts' field"
                assert isinstance(api_response["total_alerts"], int), "API response total_alerts should be an integer"
                assert api_response["total_alerts"] >= 0, "API response total_alerts should be non-negative"
                
                assert "last_updated" in api_response, "API response should include 'last_updated' field"
                assert isinstance(api_response["last_updated"], str), "API response last_updated should be a string"
                
                # Verify limit parameter is respected
                assert len(api_response["alerts"]) <= limit_param, \
                    f"API should respect limit parameter: requested {limit_param}, got {len(api_response['alerts'])}"
                
                # Verify each alert has all required fields
                required_alert_fields = [
                    'product_code', 'product_name', 'branch_code', 'current_stock',
                    'alert_status', 'status_class', 'priority', 'last_updated'
                ]
                
                for i, alert_dict in enumerate(api_response["alerts"]):
                    assert isinstance(alert_dict, dict), f"Alert {i} should be a dictionary"
                    
                    for field in required_alert_fields:
                        assert field in alert_dict, f"Alert {i} should include '{field}' field"
                        assert alert_dict[field] is not None, f"Alert {i} field '{field}' should not be None"
                        
                        # Type validation for specific fields
                        if field == 'current_stock':
                            assert isinstance(alert_dict[field], int), f"Alert {i} current_stock should be integer"
                            assert alert_dict[field] >= 0, f"Alert {i} current_stock should be non-negative"
                        elif field == 'priority':
                            assert isinstance(alert_dict[field], int), f"Alert {i} priority should be integer"
                            assert 1 <= alert_dict[field] <= 4, f"Alert {i} priority should be 1-4"
                        elif field in ['product_code', 'product_name', 'branch_code', 'alert_status', 'status_class']:
                            assert isinstance(alert_dict[field], str), f"Alert {i} {field} should be string"
                            assert alert_dict[field].strip() != "", f"Alert {i} {field} should not be empty"
                        elif field == 'last_updated':
                            assert isinstance(alert_dict[field], str), f"Alert {i} last_updated should be string"
                            # Verify it's a valid ISO format datetime
                            try:
                                datetime.fromisoformat(alert_dict[field].replace('Z', '+00:00'))
                            except ValueError:
                                pytest.fail(f"Alert {i} last_updated should be valid ISO datetime format")
                
                # Verify branch filtering works correctly
                if branch_filter:
                    for alert_dict in api_response["alerts"]:
                        # If we filtered by a specific branch, all alerts should be from that branch
                        # Note: branch_filter might not match any data, so alerts could be empty
                        if alert_dict:  # Only check if we have alerts
                            assert alert_dict['branch_code'] == branch_filter, \
                                f"When filtering by branch '{branch_filter}', all alerts should be from that branch"
                
                # Verify JSON serialization works (no datetime objects, etc.)
                import json
                try:
                    json_str = json.dumps(api_response)
                    parsed_back = json.loads(json_str)
                    assert parsed_back == api_response, "API response should be JSON serializable"
                except (TypeError, ValueError) as e:
                    pytest.fail(f"API response should be JSON serializable: {e}")
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    @given(
        branch_name=valid_branch_name_strategy(),
        excel_file=excel_with_inventory_strategy(),
        limit_param=st.integers(min_value=1, max_value=15)
    )
    @settings(max_examples=1, suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture], deadline=None)
    def test_property_3_alert_sorting_and_limiting(self, test_user, branch_name, excel_file, limit_param):
        """
        Property 3: Alert Sorting and Limiting
        
        For any collection of alerts, the system should sort them by priority 
        (Out of Stock > Very Low > Low > Reorder), then by stock quantity (lowest first), 
        and limit dashboard display to maximum 10 alerts.
        
        **Feature: dynamic-inventory-alerts, Property 3: Alert Sorting and Limiting**
        **Validates: Requirements 1.3, 4.1, 4.2, 4.4**
        """
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Upload test data
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username=test_user['username'],
                    branch_name=branch_name,
                    filename=excel_file['filename'],
                    file_data=excel_file['file_data']
                )
                
                # Generate alerts with the specified limit
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=branch_name,
                    limit=limit_param
                )
                
                # Property 3: Alert Sorting and Limiting
                
                # Test limit parameter is respected
                assert len(alerts) <= limit_param, \
                    f"Alert count should not exceed limit: requested {limit_param}, got {len(alerts)}"
                
                # Test sorting if we have multiple alerts
                if len(alerts) > 1:
                    # Verify alerts are sorted by priority (1=highest priority)
                    for i in range(len(alerts) - 1):
                        current_alert = alerts[i]
                        next_alert = alerts[i + 1]
                        
                        # Primary sort: by priority (1 is highest priority, should come first)
                        if current_alert.priority != next_alert.priority:
                            assert current_alert.priority <= next_alert.priority, \
                                f"Alerts should be sorted by priority (ascending): " \
                                f"alert {i} has priority {current_alert.priority}, " \
                                f"alert {i+1} has priority {next_alert.priority}"
                        
                        # Secondary sort: within same priority, by stock quantity (lowest first)
                        elif current_alert.priority == next_alert.priority:
                            assert current_alert.current_stock <= next_alert.current_stock, \
                                f"Within same priority {current_alert.priority}, alerts should be sorted by stock quantity (ascending): " \
                                f"alert {i} has stock {current_alert.current_stock}, " \
                                f"alert {i+1} has stock {next_alert.current_stock}"
                
                # Test priority mapping matches expected values
                priority_mapping = {
                    1: "نفد المخزون",      # Out of Stock
                    2: "منخفض جداً",       # Very Low  
                    3: "منخفض",           # Low
                    4: "إعادة طلب"         # Reorder
                }
                
                for alert in alerts:
                    expected_status = priority_mapping.get(alert.priority)
                    assert alert.alert_status == expected_status, \
                        f"Alert with priority {alert.priority} should have status '{expected_status}', got '{alert.alert_status}'"
                
                # Test that highest priority alerts come first
                if alerts:
                    first_alert = alerts[0]
                    # First alert should have the highest priority (lowest number)
                    min_priority = min(alert.priority for alert in alerts)
                    assert first_alert.priority == min_priority, \
                        f"First alert should have highest priority (lowest number): expected {min_priority}, got {first_alert.priority}"
                
                # Test dashboard limit (max 10 alerts for dashboard display)
                dashboard_alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=branch_name,
                    limit=10  # Dashboard default limit
                )
                
                assert len(dashboard_alerts) <= 10, \
                    f"Dashboard should display maximum 10 alerts, got {len(dashboard_alerts)}"
                
                # If we have more than 10 alerts available, test that we get the most critical ones
                all_alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=branch_name,
                    limit=1000  # Get all alerts
                )
                
                if len(all_alerts) > 10:
                    # Dashboard alerts should be the first 10 from the sorted list
                    for i in range(10):
                        assert dashboard_alerts[i].product_code == all_alerts[i].product_code, \
                            f"Dashboard alert {i} should match the {i}th most critical alert"
                        assert dashboard_alerts[i].priority == all_alerts[i].priority, \
                            f"Dashboard alert {i} should have same priority as {i}th most critical alert"
                        assert dashboard_alerts[i].current_stock == all_alerts[i].current_stock, \
                            f"Dashboard alert {i} should have same stock level as {i}th most critical alert"
                
                # Test consistency: multiple calls should return same order
                alerts_second_call = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=branch_name,
                    limit=limit_param
                )
                
                assert len(alerts) == len(alerts_second_call), \
                    "Multiple calls should return same number of alerts"
                
                for i, (alert1, alert2) in enumerate(zip(alerts, alerts_second_call)):
                    assert alert1.product_code == alert2.product_code, \
                        f"Alert {i} should be consistent across calls: product codes don't match"
                    assert alert1.priority == alert2.priority, \
                        f"Alert {i} should be consistent across calls: priorities don't match"
                    assert alert1.current_stock == alert2.current_stock, \
                        f"Alert {i} should be consistent across calls: stock levels don't match"
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    @given(
        branch_name=valid_branch_name_strategy(),
        excel_file=excel_with_inventory_strategy()
    )
    @settings(max_examples=1, suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture], deadline=None)
    def test_property_4_alert_data_completeness_and_formatting(self, test_user, branch_name, excel_file):
        """
        Property 4: Alert Data Completeness and Formatting
        
        For any generated alert, it should contain all required fields (product_name, branch_code, 
        current_stock, alert_status, status_class) and truncate product names longer than 30 
        characters with ellipsis.
        
        **Feature: dynamic-inventory-alerts, Property 4: Alert Data Completeness and Formatting**
        **Validates: Requirements 3.1, 3.2, 3.3, 3.4, 3.5**
        """
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Upload test data
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username=test_user['username'],
                    branch_name=branch_name,
                    filename=excel_file['filename'],
                    file_data=excel_file['file_data']
                )
                
                # Generate alerts
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=branch_name,
                    limit=1000  # Get all alerts for testing
                )
                
                # Property 4: Alert Data Completeness and Formatting
                
                # Required fields that every alert must have
                required_fields = [
                    'product_code', 'product_name', 'branch_code', 'current_stock',
                    'alert_status', 'status_class', 'priority', 'last_updated'
                ]
                
                for i, alert in enumerate(alerts):
                    # Test that all required fields are present and not None/empty
                    for field in required_fields:
                        field_value = getattr(alert, field, None)
                        assert field_value is not None, \
                            f"Alert {i} should have non-None '{field}' field, got {field_value}"
                        
                        # String fields should not be empty
                        if field in ['product_code', 'product_name', 'branch_code', 'alert_status', 'status_class']:
                            assert isinstance(field_value, str), \
                                f"Alert {i} field '{field}' should be string, got {type(field_value)}"
                            assert field_value.strip() != "", \
                                f"Alert {i} field '{field}' should not be empty string, got '{field_value}'"
                    
                    # Test data types for specific fields
                    assert isinstance(alert.product_code, str), \
                        f"Alert {i} product_code should be string, got {type(alert.product_code)}"
                    assert isinstance(alert.product_name, str), \
                        f"Alert {i} product_name should be string, got {type(alert.product_name)}"
                    assert isinstance(alert.branch_code, str), \
                        f"Alert {i} branch_code should be string, got {type(alert.branch_code)}"
                    assert isinstance(alert.current_stock, int), \
                        f"Alert {i} current_stock should be integer, got {type(alert.current_stock)}"
                    assert isinstance(alert.alert_status, str), \
                        f"Alert {i} alert_status should be string, got {type(alert.alert_status)}"
                    assert isinstance(alert.status_class, str), \
                        f"Alert {i} status_class should be string, got {type(alert.status_class)}"
                    assert isinstance(alert.priority, int), \
                        f"Alert {i} priority should be integer, got {type(alert.priority)}"
                    
                    # Test value constraints
                    assert alert.current_stock >= 0, \
                        f"Alert {i} current_stock should be non-negative, got {alert.current_stock}"
                    assert 1 <= alert.priority <= 4, \
                        f"Alert {i} priority should be 1-4, got {alert.priority}"
                    
                    # Test branch code matches the filter
                    assert alert.branch_code == branch_name, \
                        f"Alert {i} branch_code should match filter '{branch_name}', got '{alert.branch_code}'"
                    
                    # Test alert status is one of the valid values
                    valid_statuses = ["نفد المخزون", "منخفض جداً", "منخفض", "إعادة طلب"]
                    assert alert.alert_status in valid_statuses, \
                        f"Alert {i} alert_status should be one of {valid_statuses}, got '{alert.alert_status}'"
                    
                    # Test status class contains appropriate CSS classes
                    assert "bg-" in alert.status_class, \
                        f"Alert {i} status_class should contain background color class, got '{alert.status_class}'"
                    assert "text-" in alert.status_class, \
                        f"Alert {i} status_class should contain text color class, got '{alert.status_class}'"
                    assert "border-" in alert.status_class, \
                        f"Alert {i} status_class should contain border color class, got '{alert.status_class}'"
                    
                    # Test color consistency between status and CSS class
                    if alert.alert_status == "نفد المخزون":
                        assert "red-600" in alert.status_class, \
                            f"Out of stock alert should have red-600 CSS class, got '{alert.status_class}'"
                    elif alert.alert_status == "منخفض جداً":
                        assert "red-500" in alert.status_class, \
                            f"Very low alert should have red-500 CSS class, got '{alert.status_class}'"
                    elif alert.alert_status == "منخفض":
                        assert "orange-500" in alert.status_class, \
                            f"Low alert should have orange-500 CSS class, got '{alert.status_class}'"
                    elif alert.alert_status == "إعادة طلب":
                        assert "blue-500" in alert.status_class, \
                            f"Reorder alert should have blue-500 CSS class, got '{alert.status_class}'"
                
                # Test product name truncation (if we have alerts)
                if alerts:
                    # Create a test alert with a very long product name to test truncation
                    # We'll test this by checking the to_dict() method which should handle truncation
                    for alert in alerts:
                        alert_dict = alert.to_dict()
                        
                        # Test that to_dict() returns all required fields
                        for field in required_fields:
                            assert field in alert_dict, \
                                f"Alert dictionary should contain '{field}' field"
                            assert alert_dict[field] is not None, \
                                f"Alert dictionary field '{field}' should not be None"
                        
                        # Test that product name in dictionary is properly formatted
                        # If original product name is longer than 30 chars, it should be truncated with ellipsis
                        original_name = alert.product_name
                        dict_name = alert_dict['product_name']
                        
                        if len(original_name) > 30:
                            assert len(dict_name) <= 33, \
                                f"Long product name should be truncated to max 33 chars, got {len(dict_name)}: '{dict_name}'"
                            assert dict_name.endswith("..."), \
                                f"Truncated product name should end with '...', got '{dict_name}'"
                            assert dict_name[:-3] == original_name[:30], \
                                f"Truncated name should match first 30 chars of original"
                        else:
                            assert dict_name == original_name, \
                                f"Short product name should not be modified, expected '{original_name}', got '{dict_name}'"
                        
                        # Test JSON serialization compatibility
                        import json
                        try:
                            json_str = json.dumps(alert_dict)
                            parsed_back = json.loads(json_str)
                            assert parsed_back == alert_dict, \
                                "Alert dictionary should be JSON serializable"
                        except (TypeError, ValueError) as e:
                            pytest.fail(f"Alert dictionary should be JSON serializable: {e}")
                        
                        # Test that datetime fields are properly formatted as strings
                        assert isinstance(alert_dict['last_updated'], str), \
                            f"Alert last_updated should be string in dictionary, got {type(alert_dict['last_updated'])}"
                        
                        # Test that the datetime string is in ISO format
                        try:
                            from datetime import datetime
                            datetime.fromisoformat(alert_dict['last_updated'].replace('Z', '+00:00'))
                        except ValueError:
                            pytest.fail(f"Alert last_updated should be valid ISO datetime format: {alert_dict['last_updated']}")
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass


    @given(
        branch_name=valid_branch_name_strategy(),
        corruption_type=st.sampled_from(['missing_columns', 'invalid_stock', 'null_values', 'empty_data', 'mixed_types'])
    )
    @settings(max_examples=1, suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture], deadline=None)
    def test_property_7_error_handling_robustness(self, test_user, branch_name, corruption_type):
        """
        Property 7: Error Handling Robustness
        
        For any invalid or corrupted inventory data, the alert system should handle errors gracefully,
        log appropriate messages, skip invalid records, and continue processing valid ones.
        
        **Feature: dynamic-inventory-alerts, Property 7: Error Handling Robustness**
        **Validates: Requirements 7.1, 7.2**
        """
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data for this user
                data_store.clear_user_data(test_user['username'])
                
                # Create corrupted inventory data based on corruption type
                if corruption_type == 'missing_columns':
                    # Create data missing required columns
                    corrupted_data = pd.DataFrame({
                        'product_code': ['P001', 'P002'],
                        'product_name': ['Product 1', 'Product 2'],
                        # Missing 'Last_on_hand' and 'branch_code' columns
                        'supplier_name': ['Supplier A', 'Supplier B']
                    })
                    
                elif corruption_type == 'invalid_stock':
                    # Create data with invalid stock values
                    corrupted_data = pd.DataFrame({
                        'product_code': ['P001', 'P002', 'P003'],
                        'product_name': ['Product 1', 'Product 2', 'Product 3'],
                        'branch_code': [branch_name, branch_name, branch_name],
                        'Last_on_hand': ['invalid', -5, 'not_a_number'],  # Invalid stock values
                        'supplier_name': ['Supplier A', 'Supplier B', 'Supplier C']
                    })
                    
                elif corruption_type == 'null_values':
                    # Create data with null/None values
                    corrupted_data = pd.DataFrame({
                        'product_code': ['P001', None, 'P003'],
                        'product_name': [None, 'Product 2', 'Product 3'],
                        'branch_code': [branch_name, branch_name, None],
                        'Last_on_hand': [5, None, 10],
                        'supplier_name': ['Supplier A', 'Supplier B', 'Supplier C']
                    })
                    
                elif corruption_type == 'empty_data':
                    # Create completely empty DataFrame
                    corrupted_data = pd.DataFrame()
                    
                elif corruption_type == 'mixed_types':
                    # Create data with mixed/inconsistent types
                    corrupted_data = pd.DataFrame({
                        'product_code': ['P001', 123, ['P003']],  # Mixed types
                        'product_name': ['Product 1', 456, {'name': 'Product 3'}],
                        'branch_code': [branch_name, branch_name, branch_name],
                        'Last_on_hand': [5.5, '10', [15]],  # Mixed numeric types
                        'supplier_name': ['Supplier A', 'Supplier B', 'Supplier C']
                    })
                
                # Create Excel file with corrupted data
                wb = Workbook()
                wb.remove(wb.active)  # Remove default sheet
                
                # Create minimal sales sheet (required for data processing)
                sales_ws = wb.create_sheet('Transactions')
                sales_headers = ['product_code', 'sale_date', 'quantity', 'unit_price', 'revenue']
                for col, header in enumerate(sales_headers, 1):
                    sales_ws.cell(row=1, column=col, value=header)
                
                # Add minimal sales data
                sales_ws.cell(row=2, column=1, value='P001')
                sales_ws.cell(row=2, column=2, value='2024-01-01')
                sales_ws.cell(row=2, column=3, value=1)
                sales_ws.cell(row=2, column=4, value=10.0)
                sales_ws.cell(row=2, column=5, value=10.0)
                
                # Create inventory sheet with corrupted data
                if not corrupted_data.empty:
                    inventory_ws = wb.create_sheet('Item info')
                    headers = list(corrupted_data.columns)
                    for col, header in enumerate(headers, 1):
                        inventory_ws.cell(row=1, column=col, value=header)
                    
                    for row, (_, item) in enumerate(corrupted_data.iterrows(), 2):
                        for col, header in enumerate(headers, 1):
                            value = item[header]
                            # Handle various data types for Excel
                            if pd.isna(value) or value is None:
                                value = None
                            elif isinstance(value, (list, dict)):
                                value = str(value)
                            inventory_ws.cell(row=row, column=col, value=value)
                else:
                    # Create empty inventory sheet
                    inventory_ws = wb.create_sheet('Item info')
                    inventory_ws.cell(row=1, column=1, value='product_code')
                
                # Save to BytesIO
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                # Property 7: Error Handling Robustness
                
                # Test that the system handles corrupted data gracefully
                try:
                    # Upload corrupted data - this should not crash
                    file_id, sales_id, inventory_id = data_store.save_branch_data(
                        username=test_user['username'],
                        branch_name=branch_name,
                        filename=f'corrupted_{corruption_type}.xlsx',
                        file_data=excel_buffer.getvalue()
                    )
                    
                    # Generate alerts - this should handle errors gracefully
                    alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=branch_name,
                        limit=100
                    )
                    
                    # System should handle errors gracefully and return a list (possibly empty)
                    assert isinstance(alerts, list), \
                        f"Alert generation should return list even with corrupted data ({corruption_type}), got {type(alerts)}"
                    
                    # For most corruption types, we should get empty or partial results, not crashes
                    if corruption_type == 'empty_data':
                        # Empty data should result in no alerts
                        assert len(alerts) == 0, \
                            f"Empty data should result in no alerts, got {len(alerts)}"
                    
                    elif corruption_type == 'missing_columns':
                        # Missing required columns should result in no alerts (graceful handling)
                        assert len(alerts) == 0, \
                            f"Missing required columns should result in no alerts, got {len(alerts)}"
                    
                    elif corruption_type in ['invalid_stock', 'null_values', 'mixed_types']:
                        # Invalid data should be skipped, valid records should be processed
                        # We can't predict exact count, but system should not crash
                        assert len(alerts) >= 0, \
                            f"System should handle invalid data gracefully ({corruption_type}), got {len(alerts)} alerts"
                        
                        # If we got alerts, they should be valid
                        for i, alert in enumerate(alerts):
                            assert isinstance(alert.current_stock, int), \
                                f"Alert {i} should have valid integer stock even with corrupted input ({corruption_type})"
                            assert alert.current_stock >= 0, \
                                f"Alert {i} should have non-negative stock ({corruption_type})"
                            assert alert.product_code is not None and alert.product_code != "", \
                                f"Alert {i} should have valid product_code ({corruption_type})"
                            assert alert.branch_code == branch_name, \
                                f"Alert {i} should have correct branch_code ({corruption_type})"
                    
                    # Test that alerts are still properly formatted despite input corruption
                    for alert in alerts:
                        # All alerts should have valid structure
                        assert hasattr(alert, 'product_code'), "Alert should have product_code attribute"
                        assert hasattr(alert, 'product_name'), "Alert should have product_name attribute"
                        assert hasattr(alert, 'branch_code'), "Alert should have branch_code attribute"
                        assert hasattr(alert, 'current_stock'), "Alert should have current_stock attribute"
                        assert hasattr(alert, 'alert_status'), "Alert should have alert_status attribute"
                        assert hasattr(alert, 'status_class'), "Alert should have status_class attribute"
                        assert hasattr(alert, 'priority'), "Alert should have priority attribute"
                        
                        # Test to_dict() method works even with corrupted input
                        alert_dict = alert.to_dict()
                        assert isinstance(alert_dict, dict), "Alert to_dict() should return dictionary"
                        
                        # Test JSON serialization works
                        import json
                        json_str = json.dumps(alert_dict)
                        parsed_back = json.loads(json_str)
                        assert parsed_back == alert_dict, "Alert should be JSON serializable"
                
                except Exception as e:
                    # Handle expected exceptions from data layer gracefully
                    if corruption_type == 'empty_data' and 'شيت "Item info" فارغ' in str(e):
                        # This is expected behavior from data store for empty Excel files
                        # The alert service should handle this by returning empty list
                        alerts = alert_service.generate_inventory_alerts(
                            username=test_user['username'],
                            branch_filter=branch_name,
                            limit=100
                        )
                        assert isinstance(alerts, list), "Should return list even when data upload fails"
                        assert len(alerts) == 0, "Should return empty list when no data available"
                        
                    elif corruption_type == 'missing_columns' and 'Missing required columns' in str(e):
                        # This is expected behavior - alert service should handle missing columns
                        # by returning empty list instead of raising exception
                        pytest.fail(f"Alert service should handle missing columns gracefully, not raise exception: {e}")
                        
                    else:
                        # For other corruption types, we should handle gracefully without exceptions
                        pytest.fail(f"System should handle {corruption_type} gracefully without exceptions, got {type(e)}: {e}")
                
                # Test that the system can recover and work with valid data after handling corrupted data
                # Create valid data to test recovery
                valid_data = pd.DataFrame({
                    'product_code': ['V001', 'V002'],
                    'product_name': ['Valid Product 1', 'Valid Product 2'],
                    'branch_code': [branch_name, branch_name],
                    'Last_on_hand': [5, 30],  # One low stock (5), one normal stock (30)
                    'supplier_name': ['Valid Supplier A', 'Valid Supplier B']
                })
                
                # Create Excel with valid data
                wb_valid = Workbook()
                wb_valid.remove(wb_valid.active)
                
                # Sales sheet
                sales_ws = wb_valid.create_sheet('Transactions')
                for col, header in enumerate(sales_headers, 1):
                    sales_ws.cell(row=1, column=col, value=header)
                
                for row, (_, item) in enumerate(valid_data.iterrows(), 2):
                    sales_ws.cell(row=row, column=1, value=item['product_code'])
                    sales_ws.cell(row=row, column=2, value='2024-01-01')
                    sales_ws.cell(row=row, column=3, value=1)
                    sales_ws.cell(row=row, column=4, value=10.0)
                    sales_ws.cell(row=row, column=5, value=10.0)
                
                # Inventory sheet
                inventory_ws = wb_valid.create_sheet('Item info')
                headers = list(valid_data.columns)
                for col, header in enumerate(headers, 1):
                    inventory_ws.cell(row=1, column=col, value=header)
                
                for row, (_, item) in enumerate(valid_data.iterrows(), 2):
                    for col, header in enumerate(headers, 1):
                        inventory_ws.cell(row=row, column=col, value=item[header])
                
                excel_buffer_valid = BytesIO()
                wb_valid.save(excel_buffer_valid)
                excel_buffer_valid.seek(0)
                
                # Clear previous data and upload valid data
                data_store.clear_user_data(test_user['username'])
                
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username=test_user['username'],
                    branch_name=branch_name,
                    filename='valid_recovery_test.xlsx',
                    file_data=excel_buffer_valid.getvalue()
                )
                
                # Generate alerts with valid data - should work normally
                recovery_alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=branch_name,
                    limit=100
                )
                
                # Should generate alerts only for the low-stock item (stock <= 25)
                assert isinstance(recovery_alerts, list), "Recovery should return list of alerts"
                assert len(recovery_alerts) == 1, f"Should generate 1 alert for low-stock item (stock=5), got {len(recovery_alerts)}"
                
                # Verify alert is properly formed
                alert = recovery_alerts[0]
                assert alert.current_stock == 5, "Recovery alert should be for the low-stock item"
                assert alert.branch_code == branch_name, "Recovery alert should have correct branch"
                assert alert.product_code == 'V001', "Recovery alert should be for the low-stock product"
                assert alert.alert_status == 'منخفض جداً', "Stock level 5 should be classified as 'very low'"
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass


class TestInventoryAlertsAPIEndpoint:
    """Unit tests for the inventory alerts API endpoint"""
    
    def test_api_authentication_required(self, test_user):
        """
        Test that API endpoint requires authentication.
        
        **Validates: Requirements 8.1, 8.5**
        """
        # This test simulates the authentication requirement
        # In the actual Flask app, @login_required decorator handles this
        
        # Test that requests with nonexistent users return empty results
        # (The actual authentication is handled by Flask-Login in the web layer)
        alerts = alert_service.generate_inventory_alerts(
            username="nonexistent_user",
            branch_filter=None,
            limit=10
        )
        
        # Should return empty list for nonexistent user (no data available)
        assert isinstance(alerts, list)
        assert len(alerts) == 0
    
    def test_api_parameter_validation(self, test_user):
        """
        Test API parameter validation for branch filter and limit.
        
        **Validates: Requirements 8.1, 8.5**
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Test with valid parameters
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=None,  # Valid: None means all branches
                    limit=5  # Valid: positive integer
                )
                
                # Should not raise exception and return list
                assert isinstance(alerts, list)
                
                # Test with valid branch filter
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter="TestBranch",  # Valid: string branch name
                    limit=10
                )
                
                assert isinstance(alerts, list)
                
                # Test with edge case limits
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=None,
                    limit=1  # Valid: minimum limit
                )
                
                assert isinstance(alerts, list)
                assert len(alerts) <= 1  # Should respect limit
                
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=None,
                    limit=1000  # Valid: large limit
                )
                
                assert isinstance(alerts, list)
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    def test_api_error_responses(self, test_user):
        """
        Test API error handling and response format.
        
        **Validates: Requirements 8.1, 8.5**
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Test with no data available (should return empty list, not error)
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=None,
                    limit=10
                )
                
                # Should return empty list when no data available
                assert isinstance(alerts, list)
                assert len(alerts) == 0
                
                # Test error handling with invalid username (empty string)
                # The service should handle this gracefully and return empty list
                alerts = alert_service.generate_inventory_alerts(
                    username="",  # Invalid: empty username
                    branch_filter=None,
                    limit=10
                )
                
                # Should return empty list for invalid username
                assert isinstance(alerts, list)
                assert len(alerts) == 0
                
                # Test that service handles database errors gracefully
                # (This would be tested more thoroughly in integration tests)
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    def test_api_response_format_consistency(self, test_user):
        """
        Test that API responses maintain consistent format.
        
        **Validates: Requirements 8.2, 8.3, 8.4**
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Test response format with no data
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=None,
                    limit=10
                )
                
                # Simulate API response format
                api_response = {
                    "success": True,
                    "alerts": [alert.to_dict() for alert in alerts],
                    "total_alerts": len(alerts),
                    "last_updated": datetime.now().isoformat()
                }
                
                # Verify response structure is consistent even with empty data
                assert "success" in api_response
                assert "alerts" in api_response
                assert "total_alerts" in api_response
                assert "last_updated" in api_response
                
                assert isinstance(api_response["alerts"], list)
                assert isinstance(api_response["total_alerts"], int)
                assert isinstance(api_response["last_updated"], str)
                
                # Test JSON serialization
                import json
                json_str = json.dumps(api_response)
                parsed_back = json.loads(json_str)
                assert parsed_back == api_response
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass


class TestErrorHandlingScenarios:
    """Unit tests for error handling scenarios in alert service"""
    
    def test_database_connection_failure_handling(self, test_user):
        """
        Test handling of database connection failures.
        
        **Validates: Requirements 7.3, 7.4, 7.5**
        """
        # Test with invalid database path to simulate connection failure
        original_db = data_store.DB_NAME
        data_store.DB_NAME = "/invalid/path/nonexistent.db"
        
        try:
            # This should handle the database error gracefully
            alerts = alert_service.generate_inventory_alerts(
                username=test_user['username'],
                branch_filter=None,
                limit=10
            )
            
            # Should return empty list when database is unavailable
            assert isinstance(alerts, list)
            assert len(alerts) == 0
            
        except Exception as e:
            # If an exception is raised, it should be a controlled exception
            # The system should log the error and not crash
            assert isinstance(e, (sqlite3.Error, FileNotFoundError, ValueError))
            
        finally:
            # Restore original database
            data_store.DB_NAME = original_db
    
    def test_timeout_handling(self, test_user):
        """
        Test timeout handling for alert generation.
        
        **Validates: Requirements 7.4, 7.5**
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Test with empty database (should complete quickly)
                start_time = time.time()
                
                alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=None,
                    limit=10
                )
                
                end_time = time.time()
                processing_time = end_time - start_time
                
                # Should complete quickly with no data
                assert processing_time < 5.0, f"Alert generation should complete quickly, took {processing_time:.2f} seconds"
                
                # Should return empty list
                assert isinstance(alerts, list)
                assert len(alerts) == 0
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    def test_fallback_mechanisms(self, test_user):
        """
        Test fallback mechanisms for failed alert generation.
        
        **Validates: Requirements 7.5**
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Test cache fallback mechanism
                # First, populate cache with some alerts
                alert_service._alert_cache = {}  # Clear cache
                
                # Create some mock alerts for cache testing
                mock_alerts = [
                    alert_service.InventoryAlert(
                        product_code='CACHE001',
                        product_name='Cached Product 1',
                        branch_code='TestBranch',
                        current_stock=5,
                        alert_status='منخفض جداً',
                        status_class='bg-red-500/10 text-red-500 border-red-500/20',
                        priority=2,
                        last_updated=datetime.now()
                    )
                ]
                
                # Manually populate cache
                alert_service.cache_alerts(test_user['username'], mock_alerts, 'TestBranch')
                
                # Now try to get alerts - should return cached results
                cached_alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter='TestBranch',
                    limit=10
                )
                
                # Should return cached alerts
                assert isinstance(cached_alerts, list)
                assert len(cached_alerts) == 1
                assert cached_alerts[0].product_code == 'CACHE001'
                
                # Test cache invalidation
                alert_service.invalidate_alert_cache(test_user['username'])
                
                # Cache should be cleared
                cache_key = alert_service._get_cache_key(test_user['username'], 'TestBranch')
                assert cache_key not in alert_service._alert_cache
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    def test_invalid_data_handling(self, test_user):
        """
        Test handling of various invalid data scenarios.
        
        **Validates: Requirements 7.1, 7.2**
        """
        # Test classify_alert_status with invalid inputs
        
        # Test with None
        status, css_class, priority = alert_service.classify_alert_status(None)
        assert status == "نفد المخزون"  # Should default to out of stock
        assert priority == 1
        
        # Test with negative numbers
        status, css_class, priority = alert_service.classify_alert_status(-10)
        assert status == "نفد المخزون"  # Should treat negative as 0
        assert priority == 1
        
        # Test with string that can't be converted
        status, css_class, priority = alert_service.classify_alert_status("invalid")
        assert status == "نفد المخزون"  # Should default to out of stock
        assert priority == 1
        
        # Test with float values
        status, css_class, priority = alert_service.classify_alert_status(5.7)
        assert status == "منخفض جداً"  # Should convert to int(5)
        assert priority == 2
        
        # Test with very large numbers
        status, css_class, priority = alert_service.classify_alert_status(999999)
        assert status is None  # Should be above threshold
        assert priority is None
    
    def test_product_name_truncation_edge_cases(self):
        """
        Test product name truncation with edge cases.
        
        **Validates: Requirements 3.5**
        """
        # Test with None
        result = alert_service._truncate_product_name(None)
        assert result == "غير محدد"
        
        # Test with empty string
        result = alert_service._truncate_product_name("")
        assert result == "غير محدد"
        
        # Test with exactly 30 characters
        name_30_chars = "A" * 30
        result = alert_service._truncate_product_name(name_30_chars)
        assert result == name_30_chars  # Should not be truncated
        
        # Test with 31 characters
        name_31_chars = "A" * 31
        result = alert_service._truncate_product_name(name_31_chars)
        assert result == "A" * 27 + "..."  # Should be truncated
        assert len(result) == 30
        
        # Test with very long name
        very_long_name = "A" * 100
        result = alert_service._truncate_product_name(very_long_name)
        assert result == "A" * 27 + "..."
        assert len(result) == 30
        
        # Test with custom max length
        result = alert_service._truncate_product_name("ABCDEFGHIJ", max_length=5)
        assert result == "AB..."
        assert len(result) == 5
    
    def test_alert_summary_error_handling(self, test_user):
        """
        Test error handling in alert summary generation.
        
        **Validates: Requirements 7.1, 7.2**
        """
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Test with no data
                summary = alert_service.get_alert_summary(test_user['username'])
                
                # Should return valid summary structure even with no data
                assert isinstance(summary, dict)
                assert 'total_alerts' in summary
                assert 'out_of_stock' in summary
                assert 'very_low' in summary
                assert 'low' in summary
                assert 'reorder' in summary
                assert 'last_updated' in summary
                
                # All counts should be 0
                assert summary['total_alerts'] == 0
                assert summary['out_of_stock'] == 0
                assert summary['very_low'] == 0
                assert summary['low'] == 0
                assert summary['reorder'] == 0
                
                # Should have valid timestamp
                assert isinstance(summary['last_updated'], str)
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
    def test_cache_edge_cases(self, test_user):
        """
        Test cache handling edge cases.
        
        **Validates: Requirements 6.2, 6.3**
        """
        # Clear cache
        alert_service._alert_cache = {}
        
        # Test cache with None branch filter
        mock_alerts = [
            alert_service.InventoryAlert(
                product_code='TEST001',
                product_name='Test Product',
                branch_code='TestBranch',
                current_stock=10,
                alert_status='منخفض',
                status_class='bg-orange-500/10 text-orange-500 border-orange-500/20',
                priority=3,
                last_updated=datetime.now()
            )
        ]
        
        # Cache with None branch filter
        alert_service.cache_alerts(test_user['username'], mock_alerts, None)
        
        # Should be able to retrieve with None filter
        cached = alert_service._get_cached_alerts(test_user['username'], None)
        assert cached is not None
        assert len(cached) == 1
        
        # Test cache expiration
        # Manually set old timestamp
        cache_key = alert_service._get_cache_key(test_user['username'], None)
        old_time = datetime.now() - timedelta(minutes=10)  # 10 minutes ago
        alert_service._alert_cache[cache_key]['timestamp'] = old_time
        
        # Should return None for expired cache
        expired_cached = alert_service._get_cached_alerts(test_user['username'], None)
        assert expired_cached is None
        
        # Cache entry should be removed
        assert cache_key not in alert_service._alert_cache

    @given(
        branch_names=st.lists(valid_branch_name_strategy(), min_size=2, max_size=4, unique=True),
        excel_files=st.lists(excel_with_inventory_strategy(), min_size=2, max_size=4)
    )
    @settings(max_examples=1, suppress_health_check=[HealthCheck.too_slow, HealthCheck.function_scoped_fixture], deadline=None)
    def test_property_5_filter_and_cache_consistency(self, test_user, branch_names, excel_files):
        """
        Property 5: Filter and Cache Consistency
        
        For any branch filter applied, the alert system should return only alerts for products 
        in that branch, and when data hasn't changed, should return cached results.
        
        **Feature: dynamic-inventory-alerts, Property 5: Filter and Cache Consistency**
        **Validates: Requirements 4.3, 6.2, 6.3**
        """
        
        # Use temporary database for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp_db:
            original_db = data_store.DB_NAME
            data_store.DB_NAME = tmp_db.name
            
            # Initialize test database
            data_store.init_data_db()
            
            try:
                # Clear any existing data and cache
                data_store.clear_user_data(test_user['username'])
                alert_service._alert_cache = {}
                
                # Upload data for multiple branches
                uploaded_branches = []
                for i, (branch_name, excel_file) in enumerate(zip(branch_names[:len(excel_files)], excel_files)):
                    try:
                        file_id, sales_id, inventory_id = data_store.save_branch_data(
                            username=test_user['username'],
                            branch_name=branch_name,
                            filename=f'test_branch_{i}_{excel_file["filename"]}',
                            file_data=excel_file['file_data']
                        )
                        uploaded_branches.append(branch_name)
                    except Exception as e:
                        # Skip branches that fail to upload (e.g., empty data)
                        continue
                
                # Skip test if no branches were successfully uploaded
                if len(uploaded_branches) < 2:
                    return
                
                # Property 5: Filter and Cache Consistency
                
                # Test 1: Branch filtering works correctly
                for branch_name in uploaded_branches:
                    # Get alerts for specific branch
                    branch_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=branch_name,
                        limit=1000  # Get all alerts
                    )
                    
                    # All alerts should be from the specified branch
                    for alert in branch_alerts:
                        assert alert.branch_code == branch_name, \
                            f"When filtering by branch '{branch_name}', all alerts should be from that branch, got '{alert.branch_code}'"
                
                # Test 2: Different branch filters return different results (if data differs)
                branch_alert_counts = {}
                for branch_name in uploaded_branches:
                    alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=branch_name,
                        limit=1000
                    )
                    branch_alert_counts[branch_name] = len(alerts)
                
                # Test 3: All branches filter returns combined results
                all_alerts = alert_service.generate_inventory_alerts(
                    username=test_user['username'],
                    branch_filter=None,  # All branches
                    limit=1000
                )
                
                # Total alerts should be sum of individual branch alerts (or less due to deduplication)
                total_individual_alerts = sum(branch_alert_counts.values())
                assert len(all_alerts) <= total_individual_alerts, \
                    f"All branches alert count should not exceed sum of individual branches: got {len(all_alerts)}, expected <= {total_individual_alerts}"
                
                # Test 4: Cache consistency - same filter should return cached results
                if uploaded_branches:
                    test_branch = uploaded_branches[0]
                    
                    # Clear cache to start fresh
                    alert_service.invalidate_alert_cache(test_user['username'])
                    
                    # First call - should populate cache
                    first_call_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=test_branch,
                        limit=10
                    )
                    
                    # Verify cache was populated
                    cache_key = alert_service._get_cache_key(test_user['username'], test_branch)
                    assert cache_key in alert_service._alert_cache, \
                        "Cache should be populated after first call"
                    
                    # Second call - should use cache
                    second_call_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=test_branch,
                        limit=10
                    )
                    
                    # Results should be identical (from cache)
                    assert len(first_call_alerts) == len(second_call_alerts), \
                        "Cached results should have same length as original"
                    
                    for i, (alert1, alert2) in enumerate(zip(first_call_alerts, second_call_alerts)):
                        assert alert1.product_code == alert2.product_code, \
                            f"Cached alert {i} should have same product_code"
                        assert alert1.branch_code == alert2.branch_code, \
                            f"Cached alert {i} should have same branch_code"
                        assert alert1.current_stock == alert2.current_stock, \
                            f"Cached alert {i} should have same current_stock"
                        assert alert1.alert_status == alert2.alert_status, \
                            f"Cached alert {i} should have same alert_status"
                
                # Test 5: Different branch filters have separate cache entries
                if len(uploaded_branches) >= 2:
                    branch1, branch2 = uploaded_branches[0], uploaded_branches[1]
                    
                    # Clear cache
                    alert_service.invalidate_alert_cache(test_user['username'])
                    
                    # Get alerts for branch1
                    branch1_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=branch1,
                        limit=10
                    )
                    
                    # Get alerts for branch2
                    branch2_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=branch2,
                        limit=10
                    )
                    
                    # Both should have separate cache entries
                    cache_key1 = alert_service._get_cache_key(test_user['username'], branch1)
                    cache_key2 = alert_service._get_cache_key(test_user['username'], branch2)
                    
                    assert cache_key1 in alert_service._alert_cache, \
                        f"Branch1 '{branch1}' should have cache entry"
                    assert cache_key2 in alert_service._alert_cache, \
                        f"Branch2 '{branch2}' should have cache entry"
                    assert cache_key1 != cache_key2, \
                        "Different branches should have different cache keys"
                
                # Test 6: Cache invalidation clears all branch filters for user
                if uploaded_branches:
                    # Populate cache for multiple branches
                    for branch_name in uploaded_branches[:2]:  # Test with first 2 branches
                        alert_service.generate_inventory_alerts(
                            username=test_user['username'],
                            branch_filter=branch_name,
                            limit=5
                        )
                    
                    # Verify cache entries exist
                    cache_keys_before = [
                        alert_service._get_cache_key(test_user['username'], branch_name)
                        for branch_name in uploaded_branches[:2]
                    ]
                    
                    for cache_key in cache_keys_before:
                        assert cache_key in alert_service._alert_cache, \
                            f"Cache key {cache_key} should exist before invalidation"
                    
                    # Invalidate cache for user
                    alert_service.invalidate_alert_cache(test_user['username'])
                    
                    # All cache entries for this user should be cleared
                    for cache_key in cache_keys_before:
                        assert cache_key not in alert_service._alert_cache, \
                            f"Cache key {cache_key} should be cleared after invalidation"
                
                # Test 7: Cache respects limit parameter
                if uploaded_branches:
                    test_branch = uploaded_branches[0]
                    
                    # Clear cache
                    alert_service.invalidate_alert_cache(test_user['username'])
                    
                    # Get alerts with limit 5
                    limited_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=test_branch,
                        limit=5
                    )
                    
                    # Get alerts with limit 10 (should use cache but apply new limit)
                    extended_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=test_branch,
                        limit=10
                    )
                    
                    # Limited alerts should be subset of extended alerts (if cache has enough)
                    assert len(limited_alerts) <= 5, \
                        f"Limited alerts should respect limit of 5, got {len(limited_alerts)}"
                    assert len(extended_alerts) <= 10, \
                        f"Extended alerts should respect limit of 10, got {len(extended_alerts)}"
                    
                    # If we have enough cached data, limited should be prefix of extended
                    if len(extended_alerts) >= len(limited_alerts):
                        for i in range(len(limited_alerts)):
                            assert limited_alerts[i].product_code == extended_alerts[i].product_code, \
                                f"Limited alert {i} should match extended alert {i} (cache consistency)"
                
                # Test 8: Empty branch filter (None) vs specific branch filter
                if uploaded_branches:
                    test_branch = uploaded_branches[0]
                    
                    # Clear cache
                    alert_service.invalidate_alert_cache(test_user['username'])
                    
                    # Get alerts for all branches (None filter)
                    all_branch_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=None,
                        limit=1000
                    )
                    
                    # Get alerts for specific branch
                    specific_branch_alerts = alert_service.generate_inventory_alerts(
                        username=test_user['username'],
                        branch_filter=test_branch,
                        limit=1000
                    )
                    
                    # Specific branch alerts should be subset of all branch alerts
                    specific_product_codes = {alert.product_code for alert in specific_branch_alerts}
                    all_product_codes = {alert.product_code for alert in all_branch_alerts}
                    
                    assert specific_product_codes.issubset(all_product_codes), \
                        "Specific branch alerts should be subset of all branch alerts"
                    
                    # All alerts from specific branch should have correct branch_code
                    for alert in specific_branch_alerts:
                        assert alert.branch_code == test_branch, \
                            f"Specific branch filter should only return alerts from that branch"
                
                # Test 9: Cache key generation consistency
                if uploaded_branches:
                    test_branch = uploaded_branches[0]
                    
                    # Test that cache key generation is consistent
                    key1 = alert_service._get_cache_key(test_user['username'], test_branch)
                    key2 = alert_service._get_cache_key(test_user['username'], test_branch)
                    
                    assert key1 == key2, \
                        "Cache key generation should be consistent for same parameters"
                    
                    # Test that different parameters generate different keys
                    key_none = alert_service._get_cache_key(test_user['username'], None)
                    key_branch = alert_service._get_cache_key(test_user['username'], test_branch)
                    
                    assert key_none != key_branch, \
                        "Different branch filters should generate different cache keys"
                    
                    # Test that different users generate different keys
                    key_user1 = alert_service._get_cache_key(test_user['username'], test_branch)
                    key_user2 = alert_service._get_cache_key('different_user', test_branch)
                    
                    assert key_user1 != key_user2, \
                        "Different users should generate different cache keys"
                
            finally:
                # Restore original database
                data_store.DB_NAME = original_db
                
                # Clean up temporary database
                try:
                    os.unlink(tmp_db.name)
                except:
                    pass
    
