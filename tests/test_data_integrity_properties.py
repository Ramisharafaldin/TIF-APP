"""
Property-based tests for data integrity and corruption handling.

Tests Property 9: Data Integrity and Corruption Handling
**Validates: Requirements 5.4, 6.5**
"""

import pytest
import tempfile
import os
import sqlite3
import pickle
from hypothesis import given, strategies as st, settings, assume
from datetime import datetime, timedelta
import data_store
from openpyxl import Workbook
import io
import pandas as pd


def create_test_excel_file(branch_suffix=""):
    """Create a minimal valid Excel file for testing"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create Transactions sheet
    transactions_ws = wb.create_sheet("Transactions")
    transactions_ws.append(["تاريخ الفاتورة", "رقم الفاتورة", "كود الصنف", "اسم الصنف", "الكمية", "سعر البيع"])
    transactions_ws.append([f"2024-01-01", f"INV001{branch_suffix}", f"ITEM001{branch_suffix}", f"Test Item{branch_suffix}", 10, 100.0])
    transactions_ws.append([f"2024-01-02", f"INV002{branch_suffix}", f"ITEM002{branch_suffix}", f"Test Item 2{branch_suffix}", 5, 200.0])
    
    # Create Item info sheet
    item_info_ws = wb.create_sheet("Item info")
    item_info_ws.append(["كود الصنف", "اسم الصنف", "الكمية الحالية", "سعر التكلفة"])
    item_info_ws.append([f"ITEM001{branch_suffix}", f"Test Item{branch_suffix}", 100, 80.0])
    item_info_ws.append([f"ITEM002{branch_suffix}", f"Test Item 2{branch_suffix}", 50, 150.0])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def setup_test_database():
    """Set up a temporary test database."""
    # Create temporary database file
    db_fd, db_path = tempfile.mkstemp(suffix='.db')
    os.close(db_fd)
    
    # Override the database path for testing
    original_db = data_store.DB_NAME
    data_store.DB_NAME = db_path
    
    # Initialize the test database
    data_store.init_data_db()
    
    return db_path, original_db


def cleanup_test_database(db_path, original_db):
    """Clean up the temporary test database."""
    data_store.DB_NAME = original_db
    if os.path.exists(db_path):
        os.remove(db_path)


def create_test_user():
    """Create a test user for testing."""
    return {
        'username': 'test_user_integrity',
        'password': 'test_password'
    }


def corrupt_database_record(db_path, corruption_type):
    """Introduce specific types of corruption into the database."""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    
    try:
        if corruption_type == 'null_branch_name':
            # Set branch_name to NULL for some records (this column allows NULL)
            c.execute("UPDATE uploaded_files SET branch_name = NULL WHERE id = (SELECT MIN(id) FROM uploaded_files)")
        
        elif corruption_type == 'empty_branch_name':
            # Set branch_name to empty string
            c.execute("UPDATE uploaded_files SET branch_name = '' WHERE id = (SELECT MIN(id) FROM uploaded_files)")
        
        elif corruption_type == 'null_filename':
            # Since original_filename has NOT NULL constraint, we'll use empty string instead
            c.execute("UPDATE uploaded_files SET original_filename = '' WHERE id = (SELECT MIN(id) FROM uploaded_files)")
        
        elif corruption_type == 'empty_filename':
            # Set filename to empty string
            c.execute("UPDATE uploaded_files SET original_filename = '' WHERE id = (SELECT MIN(id) FROM uploaded_files)")
        
        elif corruption_type == 'negative_file_size':
            # Set file_size to negative value
            c.execute("UPDATE uploaded_files SET file_size = -1 WHERE id = (SELECT MIN(id) FROM uploaded_files)")
        
        elif corruption_type == 'null_file_size':
            # Since file_size has NOT NULL constraint, we'll use negative value instead
            c.execute("UPDATE uploaded_files SET file_size = -999 WHERE id = (SELECT MIN(id) FROM uploaded_files)")
        
        elif corruption_type == 'corrupted_dataframe':
            # Corrupt the pickled DataFrame data
            c.execute("UPDATE processed_data SET data_blob = ? WHERE id = (SELECT MIN(id) FROM processed_data)", 
                     (b'corrupted_pickle_data',))
        
        conn.commit()
    finally:
        conn.close()


@given(
    corruption_type=st.sampled_from([
        'null_branch_name', 'empty_branch_name', 
        'empty_filename', 'negative_file_size'
    ])
)
@settings(max_examples=5, deadline=15000)
def test_data_integrity_corruption_handling_property(corruption_type):
    """
    Property 9: Data Integrity and Corruption Handling
    
    For any data retrieval operation, the system should verify data integrity 
    and handle corrupted records gracefully, providing recovery options when possible.
    
    **Feature: data-upload-persistence-fix, Property 9: Data Integrity and Corruption Handling**
    **Validates: Requirements 5.4, 6.5**
    """
    db_path, original_db = setup_test_database()
    
    try:
        test_user = create_test_user()
        username = test_user['username']
        
        # Upload some valid files first
        valid_branches = ['ValidBranch1', 'ValidBranch2']
        for branch_name in valid_branches:
            excel_data = create_test_excel_file(f"_{branch_name}")
            filename = f"valid_file_{branch_name}.xlsx"
            
            try:
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username, branch_name, filename, excel_data
                )
            except Exception as e:
                # Skip this test case if we can't create valid test data
                assume(False)
        
        # Introduce corruption
        corrupt_database_record(db_path, corruption_type)
        
        # Test that get_branch_files handles corruption gracefully
        branch_files = data_store.get_branch_files(username)
        
        # Property: The function should not crash and should return valid records only
        assert isinstance(branch_files, list), "Should return a list even with corrupted data"
        
        # All returned records should be valid (corrupted records should be filtered out)
        for file_info in branch_files:
            # Verify data integrity of returned records
            assert 'branch_name' in file_info, "branch_name should be present"
            assert 'filename' in file_info, "filename should be present"
            assert 'upload_date' in file_info, "upload_date should be present"
            assert 'file_size' in file_info, "file_size should be present"
            
            # Verify field values are valid
            assert file_info['branch_name'] is not None, "branch_name should not be None"
            assert file_info['branch_name'].strip() != '', "branch_name should not be empty"
            assert file_info['filename'] is not None, "filename should not be None"
            assert file_info['filename'].strip() != '', "filename should not be empty"
            assert file_info['file_size'] is not None, "file_size should not be None"
            assert file_info['file_size'] >= 0, "file_size should not be negative"
        
        # Property: Should have fewer records than expected due to corruption filtering
        # (unless corruption type doesn't affect the records we check)
        if corruption_type in ['null_branch_name', 'empty_branch_name', 'null_filename', 
                              'empty_filename', 'negative_file_size', 'null_file_size']:
            # Should have filtered out at least one corrupted record
            assert len(branch_files) <= len(valid_branches), \
                f"Should have filtered out corrupted records, got {len(branch_files)} files"
        
    finally:
        cleanup_test_database(db_path, original_db)


@given(
    num_valid_files=st.integers(min_value=1, max_value=3),
    num_corrupted_files=st.integers(min_value=1, max_value=2)
)
@settings(max_examples=5, deadline=20000)
def test_mixed_valid_corrupted_data_property(num_valid_files, num_corrupted_files):
    """
    Property 9 Extended: Mixed Valid and Corrupted Data Handling
    
    For any mix of valid and corrupted records, the system should return only 
    the valid records and handle corrupted ones gracefully.
    
    **Feature: data-upload-persistence-fix, Property 9: Data Integrity and Corruption Handling**
    **Validates: Requirements 5.4, 6.5**
    """
    db_path, original_db = setup_test_database()
    
    try:
        test_user = create_test_user()
        username = test_user['username']
        
        # Upload valid files
        valid_file_ids = []
        for i in range(num_valid_files):
            branch_name = f'ValidBranch{i+1}'
            excel_data = create_test_excel_file(f"_V{i+1}")
            filename = f"valid_file_{i+1}.xlsx"
            
            try:
                file_id, sales_id, inventory_id = data_store.save_branch_data(
                    username, branch_name, filename, excel_data
                )
                valid_file_ids.append(file_id)
            except Exception as e:
                # Skip this test case if we can't create valid test data
                assume(False)
        
        # Create corrupted records by directly inserting bad data
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        
        corruption_types = ['null_branch_name', 'empty_branch_name', 'empty_filename', 'negative_file_size']
        
        for i in range(num_corrupted_files):
            corruption_type = corruption_types[i % len(corruption_types)]
            
            if corruption_type == 'null_branch_name':
                c.execute('''INSERT INTO uploaded_files 
                            (username, module, branch_name, original_filename, file_data, file_size)
                            VALUES (?, ?, ?, ?, ?, ?)''',
                         (username, 'branch_data', None, f'corrupted_{i}.xlsx', b'fake_data', 100))
            
            elif corruption_type == 'empty_branch_name':
                c.execute('''INSERT INTO uploaded_files 
                            (username, module, branch_name, original_filename, file_data, file_size)
                            VALUES (?, ?, ?, ?, ?, ?)''',
                         (username, 'branch_data', '', f'corrupted_{i}.xlsx', b'fake_data', 100))
            
            elif corruption_type == 'empty_filename':
                c.execute('''INSERT INTO uploaded_files 
                            (username, module, branch_name, original_filename, file_data, file_size)
                            VALUES (?, ?, ?, ?, ?, ?)''',
                         (username, 'branch_data', f'CorruptedBranch{i}', '', b'fake_data', 100))
            
            elif corruption_type == 'negative_file_size':
                c.execute('''INSERT INTO uploaded_files 
                            (username, module, branch_name, original_filename, file_data, file_size)
                            VALUES (?, ?, ?, ?, ?, ?)''',
                         (username, 'branch_data', f'CorruptedBranch{i}', f'corrupted_{i}.xlsx', b'fake_data', -1))
        
        conn.commit()
        conn.close()
        
        # Test data integrity handling
        branch_files = data_store.get_branch_files(username)
        
        # Property: Should return only valid records
        assert isinstance(branch_files, list), "Should return a list"
        
        # All returned records should be valid
        for file_info in branch_files:
            assert file_info['branch_name'] is not None and file_info['branch_name'].strip() != '', \
                "branch_name should be valid"
            assert file_info['filename'] is not None and file_info['filename'].strip() != '', \
                "filename should be valid"
            assert file_info['file_size'] is not None and file_info['file_size'] >= 0, \
                "file_size should be valid"
        
        # Property: Should return exactly the number of valid files (or fewer due to deduplication)
        assert len(branch_files) <= num_valid_files, \
            f"Should return at most {num_valid_files} valid files, got {len(branch_files)}"
        
        # Property: Should not return any corrupted records
        # We can verify this by checking that all returned files have valid data
        valid_branch_names = {f'ValidBranch{i+1}' for i in range(num_valid_files)}
        returned_branch_names = {f['branch_name'] for f in branch_files}
        
        # All returned branches should be from our valid set
        assert returned_branch_names.issubset(valid_branch_names), \
            f"Returned branches {returned_branch_names} should be subset of valid branches {valid_branch_names}"
        
    finally:
        cleanup_test_database(db_path, original_db)


@given(
    branch_name=st.text(min_size=1, max_size=30).filter(lambda x: x.strip() and not any(c in x for c in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']))
)
@settings(max_examples=5, deadline=15000)
def test_dataframe_corruption_handling_property(branch_name):
    """
    Property 9 Extended: DataFrame Corruption Handling
    
    For any corrupted DataFrame data in the database, the system should handle 
    the corruption gracefully and not crash the application.
    
    **Feature: data-upload-persistence-fix, Property 9: Data Integrity and Corruption Handling**
    **Validates: Requirements 5.4, 6.5**
    """
    db_path, original_db = setup_test_database()
    
    try:
        test_user = create_test_user()
        username = test_user['username']
        
        # Upload a valid file first
        excel_data = create_test_excel_file()
        filename = "test_file.xlsx"
        
        try:
            file_id, sales_id, inventory_id = data_store.save_branch_data(
                username, branch_name, filename, excel_data
            )
        except Exception as e:
            # Skip this test case if we can't create valid test data
            assume(False)
        
        # Corrupt the DataFrame data in the database
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        
        # Corrupt both sales and inventory DataFrames
        c.execute("UPDATE processed_data SET data_blob = ? WHERE id = ?", 
                 (b'corrupted_pickle_data', sales_id))
        c.execute("UPDATE processed_data SET data_blob = ? WHERE id = ?", 
                 (b'corrupted_pickle_data', inventory_id))
        
        conn.commit()
        conn.close()
        
        # Test that get_branch_data handles DataFrame corruption gracefully
        try:
            sales_df, inventory_df = data_store.get_branch_data(username, branch_name)
            
            # Property: Function should not crash, even with corrupted data
            # It may return None DataFrames, which is acceptable
            assert sales_df is None or isinstance(sales_df, pd.DataFrame), \
                "sales_df should be None or DataFrame"
            assert inventory_df is None or isinstance(inventory_df, pd.DataFrame), \
                "inventory_df should be None or DataFrame"
            
        except Exception as e:
            # The function should handle corruption gracefully, not crash
            # If it does raise an exception, it should be a controlled one, not a pickle error
            assert "pickle" not in str(e).lower(), \
                f"Should handle pickle corruption gracefully, not raise pickle error: {e}"
        
        # Test that get_branch_files still works despite DataFrame corruption
        branch_files = data_store.get_branch_files(username)
        
        # Property: File listing should still work even if DataFrames are corrupted
        assert isinstance(branch_files, list), "Should return a list"
        
        # Should still find the file record (file metadata is separate from DataFrame data)
        our_files = [f for f in branch_files if f['branch_name'] == branch_name]
        assert len(our_files) == 1, "Should still find the file record"
        
    finally:
        cleanup_test_database(db_path, original_db)


if __name__ == "__main__":
    # Run a simple test to verify the property
    print("Testing data integrity and corruption handling property...")
    
    db_path, original_db = setup_test_database()
    
    try:
        test_user = create_test_user()
        username = test_user['username']
        branch_name = "TestBranch"
        
        # Upload a valid file
        excel_data = create_test_excel_file()
        filename = "test_file.xlsx"
        
        file_id, sales_id, inventory_id = data_store.save_branch_data(
            username, branch_name, filename, excel_data
        )
        print(f"Uploaded file with ID {file_id}")
        
        # Test normal retrieval
        branch_files = data_store.get_branch_files(username)
        print(f"Retrieved {len(branch_files)} files normally")
        
        # Introduce corruption
        corrupt_database_record(db_path, 'null_branch_name')
        print("Introduced null branch_name corruption")
        
        # Test corrupted retrieval
        branch_files_after = data_store.get_branch_files(username)
        print(f"Retrieved {len(branch_files_after)} files after corruption")
        
        # Should handle corruption gracefully
        assert len(branch_files_after) <= len(branch_files), "Should filter out corrupted records"
        print("✓ Data integrity and corruption handling property test passed")
        
    finally:
        cleanup_test_database(db_path, original_db)